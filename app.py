from flask import Flask, render_template, jsonify, request
import json
import os
import random
import re
from datetime import datetime, timedelta
import locale
import hashlib
import shutil
from deep_translator import GoogleTranslator
from deep_translator.exceptions import TranslationNotFound, NotValidPayload
import sys

# Configurar codificación UTF-8 para la consola
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

app = Flask(__name__)

# Register ChatBotAI Blueprint
from ChatBotAI import chatbot_bp
app.register_blueprint(chatbot_bp)

# Configuración
JSON_FOLDER_PATH = r"C:\n8n_Docker\Files"
APARTMENT_CONFIG_FILE = "apartment_config.json"
ID_MAPPING_FILE = r"C:\n8n_Docker\Files\ID.txt"
IDB_MAPPING_FILE = r"C:\n8n_Docker\Files\IDB.txt"

# Nombre de los archivos JSON
JSON_FILE_NAME = "DatasetScrAN.json"  # Airbnb reviews (updated data source)
BOOKING_JSON_FILE_NAME = "DatasetScrBookingAN.json"  # Booking reviews (updated data source)
GENERAL_REVIEWS_FILE_NAME = "GeneralReviews.json"  # General problems

# Configuración de snapshots
SNAPSHOTS_DIR = os.path.join(os.path.dirname(__file__), 'data', 'snapshots')
SNAPSHOTS_METADATA_FILE = os.path.join(os.path.dirname(__file__), 'data', 'snapshots_metadata.json')
LAST_HASH_FILE = os.path.join(os.path.dirname(__file__), 'data', 'last_dataset_hash.txt')

# Configuración de reviews históricas
HISTORICAL_REVIEWS_DIR = os.path.join(os.path.dirname(__file__), 'data', 'DataProblemListing')

# Máximo de días de antigüedad para mostrar reviews (30 días = 1 mes)
MAX_REVIEW_AGE_DAYS = 30

# IDs de apartamentos a excluir (no mostrar comentarios)
EXCLUDED_APARTMENT_IDS = {
    '50587278',
    '814427016412775340'
}

# Cache de traducciones para evitar llamadas repetidas
translation_cache = {}

# Cache de reviews para evitar recargas constantes
reviews_cache = {
    'neuen': {'data': None, 'timestamp': None},
    'allem': {'data': None, 'timestamp': None}
}

# Tiempo de vida del caché en segundos (5 minutos)
CACHE_TTL = 300

def normalize_probleme(wohnungen):
    """Normalize field names in probleme to handle encoding issues with erwähnungen.

    The AI workflow sometimes outputs 'erw??hnungen' instead of 'erwähnungen' due to
    character encoding issues. This function normalizes the field name to ensure
    consistent access to mention counts.
    """
    for apt in wohnungen:
        if 'probleme' in apt and isinstance(apt['probleme'], list):
            for prob in apt['probleme']:
                # Handle corrupted field name 'erw??hnungen' -> 'erwähnungen'
                if 'erw??hnungen' in prob and 'erwähnungen' not in prob:
                    prob['erwähnungen'] = prob.pop('erw??hnungen')
                # Also check for other possible encoding variations
                for key in list(prob.keys()):
                    if key.startswith('erw') and key.endswith('hnungen') and key != 'erwähnungen':
                        prob['erwähnungen'] = prob.pop(key)
    return wohnungen

def load_id_mapping():
    """Cargar mapeo de IDs desde ID.txt (para Airbnb)"""
    id_mapping = {}
    try:
        with open(ID_MAPPING_FILE, 'r', encoding='utf-8') as file:
            for line in file:
                line = line.strip()
                if not line or line.startswith('http'):
                    continue
                # Formato: "609172560881241855 (UT)"
                match = re.match(r'(\d+)\s*\(([^)]+)\)', line)
                if match:
                    apt_id = match.group(1)
                    apt_code = match.group(2)
                    id_mapping[apt_id] = apt_code
        return id_mapping
    except Exception as e:
        print(f"⚠️  Error cargando ID.txt: {e}")
        return {}

def load_idb_mapping():
    """Cargar mapeo de URLs de Booking desde IDB.txt"""
    idb_mapping = {}
    try:
        with open(IDB_MAPPING_FILE, 'r', encoding='utf-8') as file:
            for line in file:
                line = line.strip()
                if not line or not line.startswith('https://www.booking.com'):
                    continue
                # Formato: "https://www.booking.com/hotel/de/... (CODE)"
                match = re.match(r'(https://[^\s]+)\s+\(([^)]+)\)', line)
                if match:
                    booking_url = match.group(1)
                    apt_code = match.group(2)
                    idb_mapping[booking_url] = apt_code
        return idb_mapping
    except Exception as e:
        print(f"⚠️  Error cargando IDB.txt: {e}")
        return {}

def extract_apartment_id_from_url(url):
    """Extraer ID del apartamento de la URL de Airbnb"""
    if not url:
        return None
    match = re.search(r'/rooms/([0-9]+)', url)
    return match.group(1) if match else None

def load_apartment_config():
    """Cargar configuración de apartamentos desde JSON"""
    try:
        config_path = os.path.join(os.path.dirname(__file__), APARTMENT_CONFIG_FILE)
        with open(config_path, 'r', encoding='utf-8') as file:
            return json.load(file)
    except Exception as e:
        print(f"⚠️  Error cargando configuración de apartamentos: {e}")
        # Configuración de respaldo
        return {
            "apartments": {
                '649153494847068923': {'code': 'H4', 'name': 'H4 - Appartment im maritimen Stil'},
                '940339949730055972': {'code': 'F3', 'name': 'F3 - Gemütliche Ferienwohnung'},
                '609172560881241855': {'code': 'UT', 'name': 'UT - Apartment mit Balkon und Flussblick'}
            },
            "default_apartment": {'code': 'UNK', 'name': 'Unbekanntes Apartment'}
        }

def get_apartment_name_from_url(listing_url):
    """Obtener nombre del apartamento desde la URL usando ID.txt (Airbnb)"""
    apartment_id = extract_apartment_id_from_url(listing_url)
    if not apartment_id:
        return 'Wohnung'
    
    # Intentar obtener código del ID.txt
    id_mapping = load_id_mapping()
    if apartment_id in id_mapping:
        code = id_mapping[apartment_id]
        return code  # Solo el código, sin "- Wohnung"
    
    # Si no está en ID.txt, intentar con config
    config = load_apartment_config()
    apartments = config.get('apartments', {})
    if apartment_id in apartments:
        return apartments[apartment_id]['name']
    
    # Valor por defecto
    return f'Wohnung {apartment_id[:6]}...'

def get_apartment_code_from_booking_url(booking_url):
    """Obtener código del apartamento desde URL de Booking usando IDB.txt"""
    if not booking_url:
        return None
    
    # Primero, intentar extraer código si está en el formato "URL (CODE)"
    match = re.search(r'\(([^)]+)\)\s*$', booking_url)
    if match:
        return match.group(1)
    
    # Cargar mapeo de IDB.txt
    idb_mapping = load_idb_mapping()
    
    # Buscar coincidencia exacta primero
    booking_url_clean = booking_url.strip()
    if booking_url_clean in idb_mapping:
        return idb_mapping[booking_url_clean]
    
    # Buscar coincidencia parcial (sin query params o trailing slashes)
    booking_url_base = booking_url_clean.split('?')[0].rstrip('/')
    for mapped_url, code in idb_mapping.items():
        mapped_url_base = mapped_url.split('?')[0].rstrip('/')
        if booking_url_base == mapped_url_base:
            return code
    
    # Si no se encuentra, retornar None
    return None

def calculate_file_hash(file_path):
    """Calcular hash MD5 de un archivo para detectar cambios"""
    try:
        hash_md5 = hashlib.md5()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except Exception as e:
        print(f"⚠️  Error calculando hash: {e}")
        return None

def get_last_saved_hash():
    """Obtener el último hash guardado del dataset"""
    try:
        if os.path.exists(LAST_HASH_FILE):
            with open(LAST_HASH_FILE, 'r', encoding='utf-8') as f:
                return f.read().strip()
        return None
    except Exception as e:
        print(f"⚠️  Error leyendo hash guardado: {e}")
        return None

def save_hash(file_hash):
    """Guardar el hash actual del dataset"""
    try:
        os.makedirs(os.path.dirname(LAST_HASH_FILE), exist_ok=True)
        with open(LAST_HASH_FILE, 'w', encoding='utf-8') as f:
            f.write(file_hash)
        return True
    except Exception as e:
        print(f"⚠️  Error guardando hash: {e}")
        return False

def load_snapshots_metadata():
    """Cargar metadata de snapshots existentes"""
    try:
        if os.path.exists(SNAPSHOTS_METADATA_FILE):
            with open(SNAPSHOTS_METADATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {'snapshots': []}
    except Exception as e:
        print(f"⚠️  Error cargando metadata de snapshots: {e}")
        return {'snapshots': []}

def save_snapshots_metadata(metadata):
    """Guardar metadata de snapshots"""
    try:
        os.makedirs(os.path.dirname(SNAPSHOTS_METADATA_FILE), exist_ok=True)
        with open(SNAPSHOTS_METADATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"⚠️  Error guardando metadata de snapshots: {e}")
        return False

def create_snapshot(source_file_path):
    """Crear snapshot del dataset actual"""
    try:
        # Asegurar que existe el directorio de snapshots
        os.makedirs(SNAPSHOTS_DIR, exist_ok=True)
        
        # Cargar metadata existente
        metadata = load_snapshots_metadata()
        snapshots = metadata.get('snapshots', [])
        
        # Calcular siguiente número de snapshot
        next_id = len(snapshots) + 1
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        snapshot_filename = f"dataset_{next_id:03d}_{timestamp}.json"
        snapshot_path = os.path.join(SNAPSHOTS_DIR, snapshot_filename)
        
        # Copiar archivo
        shutil.copy2(source_file_path, snapshot_path)
        
        # Calcular hash del snapshot
        file_hash = calculate_file_hash(snapshot_path)
        
        # Cargar datos para obtener estadísticas básicas
        with open(snapshot_path, 'r', encoding='utf-8') as f:
            reviews_data = json.load(f)
        
        # Contar apartamentos únicos
        apartments = set()
        for review in reviews_data:
            listing_url = review.get('listingUrl', '')
            apt_id = extract_apartment_id_from_url(listing_url)
            if apt_id:
                apartments.add(apt_id)
        
        # Agregar metadata del nuevo snapshot
        snapshot_info = {
            'id': next_id,
            'filename': snapshot_filename,
            'created_at': datetime.now().isoformat(),
            'total_reviews': len(reviews_data),
            'apartments_count': len(apartments),
            'hash': file_hash
        }
        
        snapshots.append(snapshot_info)
        metadata['snapshots'] = snapshots
        
        # Guardar metadata actualizada
        save_snapshots_metadata(metadata)
        
        print(f"✅ Snapshot creado: {snapshot_filename}")
        print(f"   📊 Reviews: {len(reviews_data)} | Apartamentos: {len(apartments)}")
        
        return snapshot_info
        
    except Exception as e:
        print(f"❌ Error creando snapshot: {e}")
        import traceback
        traceback.print_exc()
        return None

def check_and_create_snapshot():
    """Verificar si el dataset cambió y crear snapshot si es necesario"""
    try:
        json_file_path = os.path.join(JSON_FOLDER_PATH, JSON_FILE_NAME)
        
        if not os.path.exists(json_file_path):
            return False
        
        # Calcular hash actual del dataset
        current_hash = calculate_file_hash(json_file_path)
        if not current_hash:
            return False
        
        # Obtener hash guardado anteriormente
        last_hash = get_last_saved_hash()
        
        # Si el hash es diferente (o es la primera vez), crear snapshot
        if current_hash != last_hash:
            print(f"\n🔍 Dataset modificado detectado (hash: {current_hash[:8]}...)")
            
            # Crear snapshot
            snapshot_info = create_snapshot(json_file_path)
            
            if snapshot_info:
                # Guardar nuevo hash
                save_hash(current_hash)
                return True
        
        return False
        
    except Exception as e:
        print(f"⚠️  Error en verificación de snapshot: {e}")
        return False

def convert_iso_to_display(iso_date_str):
    """Convertir fecha ISO a formato de display (ej: 2025-10-12T13:49:27Z -> Oktober 2025)"""
    if not iso_date_str:
        return ''
    
    try:
        # Parsear fecha ISO
        dt = datetime.fromisoformat(iso_date_str.replace('Z', '+00:00'))
        
        # Calcular diferencia en días desde ahora
        now = datetime.now(dt.tzinfo) if dt.tzinfo else datetime.now()
        days_diff = (now - dt).days
        
        # Si es de hace menos de 7 días, usar "Vor X Tag"
        if days_diff == 0:
            return "Heute"
        elif days_diff == 1:
            return "Vor 1 Tag"
        elif days_diff < 7:
            return f"Vor {days_diff} Tage"
        elif days_diff < 30:
            weeks = days_diff // 7
            if weeks == 1:
                return "Vor 1 Woche"
            else:
                return f"Vor {weeks} Wochen"
        else:
            # Usar formato "Monat Jahr"
            month_names_de = [
                'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
                'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember'
            ]
            month_name = month_names_de[dt.month - 1]
            return f"{month_name} {dt.year}"
    except Exception as e:
        print(f"⚠️  Error convirtiendo fecha '{iso_date_str}': {e}")
        return iso_date_str

def is_review_recent_iso(iso_date_str, max_days=30):
    """Verificar si una review es reciente basado en fecha ISO"""
    if not iso_date_str:
        return False
    
    try:
        # Parsear fecha ISO
        review_dt = datetime.fromisoformat(iso_date_str.replace('Z', '+00:00'))
        now = datetime.now(review_dt.tzinfo) if review_dt.tzinfo else datetime.now()
        
        # Calcular diferencia en días
        days_diff = (now - review_dt).days
        
        is_recent = days_diff <= max_days
        if days_diff <= 1:
            print(f"   Review: {days_diff} día(s) - INCLUIR")
        
        return is_recent
    except Exception as e:
        print(f"⚠️  Error parseando fecha ISO '{iso_date_str}': {e}")
        return False

def is_review_recent(date_str, max_days=30):
    """Verificar si una review es de los últimos N días (default: 30 días / 1 mes)"""
    if not date_str:
        return False  # Si no hay fecha, NO incluir la review
    
    now = datetime.now()
    date_str_lower = date_str.lower()
    
    try:
        # Formato: "Vor X Tag" o "Vor X Tage"
        if 'vor' in date_str_lower and 'tag' in date_str_lower:
            match = re.search(r'(\d+)', date_str)
            if match:
                days = int(match.group(1))
                is_recent = days <= max_days
                print(f"   Fecha '{date_str}': {days} días - {'INCLUIR' if is_recent else 'FILTRAR'}")
                return is_recent
        
        # Formato: "Vor X Woche" o "Vor X Wochen"
        if 'vor' in date_str_lower and 'woche' in date_str_lower:
            match = re.search(r'(\d+)', date_str)
            if match:
                weeks = int(match.group(1))
                days = weeks * 7
                is_recent = days <= max_days
                print(f"   Fecha '{date_str}': {weeks} semanas ({days} días) - {'INCLUIR' if is_recent else 'FILTRAR'}")
                return is_recent
        
        # Formato: "Oktober 2025", "September 2025", etc.
        month_names_de = {
            'januar': 1, 'februar': 2, 'märz': 3, 'april': 4,
            'mai': 5, 'juni': 6, 'juli': 7, 'august': 8,
            'september': 9, 'oktober': 10, 'november': 11, 'dezember': 12
        }
        
        for month_name, month_num in month_names_de.items():
            if month_name in date_str_lower:
                year_match = re.search(r'(\d{4})', date_str)
                if year_match:
                    year = int(year_match.group(1))
                    # Usar el último día del mes para ser más inclusivo
                    # Si es el mes actual, comparar desde hoy
                    if year == now.year and month_num == now.month:
                        # Es el mes actual - INCLUIR
                        print(f"   Fecha '{date_str}': Mes actual - INCLUIR")
                        return True
                    else:
                        # Calcular diferencia desde el último día del mes
                        # Para mes anterior, usar último día
                        if month_num == 12:
                            next_month = datetime(year + 1, 1, 1)
                        else:
                            next_month = datetime(year, month_num + 1, 1)
                        last_day_of_month = next_month - timedelta(days=1)
                        days_diff = (now - last_day_of_month).days
                        is_recent = days_diff <= max_days
                        print(f"   Fecha '{date_str}': {days_diff} días desde fin de mes - {'INCLUIR' if is_recent else 'FILTRAR'}")
                        return is_recent
        
        # Si no coincide con ningún formato conocido, NO incluir
        print(f"   Fecha '{date_str}': Formato desconocido - FILTRAR")
        return False
        
    except (ValueError, AttributeError) as e:
        print(f"⚠️  Error parseando fecha '{date_str}': {e}")
        return False

def translate_traveler_type(traveler_type):
    """Traducir tipo de viajero de Booking al alemán"""
    if not traveler_type:
        return None
    
    traveler_type_lower = traveler_type.lower().strip()
    
    # Diccionario de traducciones
    translations = {
        'couple': 'Paar',
        'family': 'Familie',
        'family with children': 'Familie mit Kindern',
        'friends': 'Freunde',
        'solo': 'Solo Reisende',
        'solo traveler': 'Solo Reisende',
        'business': 'Geschäftsreise',
        'business traveler': 'Geschäftsreise',
        'group': 'Gruppe',
    }
    
    # Buscar coincidencia exacta
    for en_key, de_value in translations.items():
        if en_key == traveler_type_lower:
            return de_value
    
    # Si no hay coincidencia exacta, devolver el original
    return traveler_type if traveler_type else None

def translate_room_info(room_info):
    """Traducir información de la habitación (ej: '1 room, 2 guests' -> '1 Zimmer, 2 Gäste')"""
    if not room_info:
        return None
    
    result = room_info
    
    # Hacer traducciones de frases completas primero (antes de palabras individuales)
    phrase_translations = {
        'single room': 'Einzelzimmer',
        'double room': 'Doppelzimmer',
        'twin room': 'Zweibettzimmer',
        'family room': 'Familienzimmer',
        'one-bedroom apartment': 'Einschlafzimmer-Wohnung',
        'one bedroom apartment': 'Einschlafzimmer-Wohnung',
        'two-bedroom apartment': 'Zweischlafzimmer-Wohnung',
        'two bedroom apartment': 'Zweischlafzimmer-Wohnung',
        'studio apartment': 'Studio-Wohnung',
        'one-bedroom': 'Einschlafzimmer',
        'two-bedroom': 'Zweischlafzimmer',
    }
    
    for en_phrase, de_phrase in phrase_translations.items():
        result = re.sub(r'\b' + en_phrase + r'\b', de_phrase, result, flags=re.IGNORECASE)
    
    # Luego hacer traducciones de palabras individuales
    word_translations = {
        'room': 'Zimmer',
        'rooms': 'Zimmer',
        'guest': 'Gast',
        'guests': 'Gäste',
        'bedroom': 'Schlafzimmer',
        'bedrooms': 'Schlafzimmer',
    }
    
    for en_word, de_word in word_translations.items():
        # Reemplazar palabras completas (case-insensitive)
        result = re.sub(r'\b' + en_word + r'\b', de_word, result, flags=re.IGNORECASE)
    
    return result

def translate_to_german(text, source_lang='auto'):
    """Traducir texto al alemán si no está en alemán
    
    Args:
        text: Texto a traducir
        source_lang: Código de idioma fuente ('auto' para detección automática)
    
    Returns:
        Tupla (texto_traducido, fue_traducido, idioma_original)
    """
    if not text or not isinstance(text, str) or len(text.strip()) < 3:
        return text, False, 'unknown'
    
    # Limpiar HTML tags para mejor detección de idioma
    text_clean = re.sub(r'<[^>]+>', '', text).strip()
    
    # Verificar cache primero
    cache_key = hashlib.md5(text_clean.encode('utf-8')).hexdigest()
    if cache_key in translation_cache:
        cached = translation_cache[cache_key]
        return cached['translated_text'], cached['was_translated'], cached['source_lang']
    
    try:
        # Intentar traducir usando Google Translator
        # deep-translator detecta automáticamente si es necesario traducir
        translator = GoogleTranslator(source='auto', target='de')
        translated_text = translator.translate(text_clean)
        
        # Si el texto traducido es igual al original, probablemente ya era alemán
        was_translated = translated_text != text_clean
        source_lang = 'de' if not was_translated else 'auto'
        
        # Guardar en cache
        translation_cache[cache_key] = {
            'translated_text': translated_text,
            'was_translated': was_translated,
            'source_lang': source_lang
        }
        
        return translated_text, was_translated, source_lang
        
    except (TranslationNotFound, NotValidPayload) as e:
        # Si no se puede traducir, retornar texto original
        print(f"⚠️  No se pudo traducir: {e}")
        return text, False, 'unknown'
    except Exception as e:
        # Si falla la traducción, retornar texto original
        print(f"⚠️  Error traduciendo texto: {e}")
        return text, False, 'error'

def load_reviews():
    """Cargar y procesar las reseñas desde DatasetScr.json (Airbnb) y DatasetScrBooking.json (Booking)"""
    # Verificar si hay datos en caché válidos
    now = datetime.now()
    if reviews_cache['neuen']['data'] is not None and reviews_cache['neuen']['timestamp'] is not None:
        elapsed = (now - reviews_cache['neuen']['timestamp']).total_seconds()
        if elapsed < CACHE_TTL:
            print(f"\n⚡ Usando reviews en caché (edad: {int(elapsed)}s)")
            # Crear copia y mezclar aleatoriamente para mantener aleatoriedad
            cached_reviews = reviews_cache['neuen']['data'].copy()
            random.shuffle(cached_reviews)
            return cached_reviews
    
    print("\n🔄 Cargando reviews nuevas desde archivos...")
    processed_reviews = []
    apartment_stats = {}
    reviews_filtered_by_apartment = 0
    reviews_filtered_by_date = 0
    reviews_filtered_by_empty_content = 0  # Nuevo: reviews vacías de Booking
    total_reviews_found_airbnb = 0
    total_reviews_found_booking = 0
    unique_apartments = set()
    
    try:
        # Verificar y crear snapshot si el dataset cambió
        check_and_create_snapshot()
        
        # ===== CARGAR REVIEWS DE AIRBNB =====
        # Solo cargar desde DatasetScrAN.json (archivo principal)
        airbnb_files = []
        airbnb_file = os.path.join(JSON_FOLDER_PATH, JSON_FILE_NAME)
        if os.path.exists(airbnb_file):
            airbnb_files.append(airbnb_file)
        
        reviews_data_airbnb = []
        
        if airbnb_files:
            print(f"\n📁 Cargando reviews de Airbnb: {len(airbnb_files)} archivo(s) encontrado(s)")
            
            for file_path in airbnb_files:
                try:
                    with open(file_path, 'r', encoding='utf-8') as file:
                        data = json.load(file)
                        if isinstance(data, list):
                            reviews_data_airbnb.extend(data)
                            print(f"   ✓ {os.path.basename(file_path)}: {len(data)} reviews")
                except Exception as e:
                    print(f"   ⚠️  Error leyendo {os.path.basename(file_path)}: {e}")
            
            if isinstance(reviews_data_airbnb, list):
                total_reviews_found_airbnb = len(reviews_data_airbnb)
                print(f"✅ {total_reviews_found_airbnb} reviews de Airbnb encontradas")
                
                # Procesar cada review de Airbnb
                for review in reviews_data_airbnb:
                    listing_url = review.get('listingUrl', '')
                    reviewer_name = review.get('reviewerName', 'Anonym')
                    reviewer_picture = review.get('reviewerProfilePicture', '')
                    review_date_iso = review.get('reviewDate', '')
                    review_text = review.get('reviewText', '')
                    language = review.get('language', 'de')
                    review_id_num = review.get('reviewId', 0)
                    
                    # Extraer apartment_id de la URL
                    apartment_id = extract_apartment_id_from_url(listing_url)
                    if not apartment_id:
                        continue
                    
                    unique_apartments.add(apartment_id)
                    
                    # Filtrar apartamentos excluidos
                    if apartment_id in EXCLUDED_APARTMENT_IDS:
                        reviews_filtered_by_apartment += 1
                        continue
                    
                    # Convertir fecha ISO a formato legible
                    date_display = convert_iso_to_display(review_date_iso)
                    
                    # Filtrar reviews antiguas
                    if not is_review_recent_iso(review_date_iso, max_days=MAX_REVIEW_AGE_DAYS):
                        reviews_filtered_by_date += 1
                        continue
                    
                    # Obtener nombre del apartamento
                    apartment_name = get_apartment_name_from_url(listing_url)
                    
                    # Inicializar estadísticas del apartamento
                    if apartment_id not in apartment_stats:
                        apartment_stats[apartment_id] = {'count': 0, 'name': apartment_name, 'source': 'Airbnb'}
                    
                    # Generar ID único
                    review_id = f"airbnb_{review_id_num}_{apartment_id}"
                    
                    # Procesar rating (Airbnb usa escala de 5)
                    rating = review.get('rating', 0)
                    if not isinstance(rating, int):
                        rating = int(rating) if rating else 0
                    
                    processed_review = {
                        'id': review_id,
                        'text': review_text,
                        'rating': rating,
                        'max_rating': 5,  # Escala máxima de Airbnb
                        'platform_code': 'AB',  # Siglas para Airbnb
                        'date': date_display,
                        'reviewer_name': reviewer_name,
                        'reviewer_picture': reviewer_picture,
                        'reviewer_location': '',
                        'host_name': 'Urlaubsmagie',
                        'created_at': review_date_iso,
                        'language': language,
                        'apartment_id': apartment_id,
                        'apartment_name': apartment_name,
                        'apartment_url': listing_url,
                        'source': 'Airbnb',
                        'stay_info': None  # Airbnb no tiene información de estadías
                    }
                    
                    processed_reviews.append(processed_review)
                    apartment_stats[apartment_id]['count'] += 1
        else:
            print(f"\n⚠️  Archivos Airbnb no encontrados")
        
        # ===== CARGAR REVIEWS DE BOOKING =====
        # Solo cargar desde DatasetScrBookingAN.json (archivo principal)
        booking_files = []
        booking_file = os.path.join(JSON_FOLDER_PATH, BOOKING_JSON_FILE_NAME)
        if os.path.exists(booking_file):
            booking_files.append(booking_file)
        
        reviews_data_booking = []
        
        if booking_files:
            print(f"\n📁 Cargando reviews de Booking: {len(booking_files)} archivo(s) encontrado(s)")
            
            for file_path in booking_files:
                try:
                    with open(file_path, 'r', encoding='utf-8') as file:
                        data = json.load(file)
                        if isinstance(data, list):
                            reviews_data_booking.extend(data)
                            print(f"   ✓ {os.path.basename(file_path)}: {len(data)} reviews")
                except Exception as e:
                    print(f"   ⚠️  Error leyendo {os.path.basename(file_path)}: {e}")
            
            total_reviews_found_booking = len(reviews_data_booking)
            print(f"✅ {total_reviews_found_booking} reviews de Booking encontradas")
            
            # Procesar cada review de Booking
            for review in reviews_data_booking:
                    booking_url = review.get('startUrl', '')
                    reviewer_name = review.get('userName', 'Anonym')
                    reviewer_picture = review.get('userAvatar', '')
                    review_date_iso = review.get('reviewDate', '')
                    rating_booking = review.get('rating', 0)
                    
                    # Validar que haya rating válido
                    if not rating_booking or (isinstance(rating_booking, (int, float)) and rating_booking == 0):
                        continue  # Saltar reviews sin rating
                    
                    # Construir texto de review de Booking
                    review_title = review.get('reviewTitle')
                    liked_text = review.get('likedText')
                    disliked_text = review.get('dislikedText')
                    room_info = review.get('roomInfo', '')
                    traveler_type = review.get('travelerType', '')
                    
                    # Combinar textos de review (sin traducción automática para optimizar velocidad)
                    review_parts = []
                    
                    # Agregar review_title sin traducción
                    if review_title:
                        review_parts.append(f"<b>{review_title}</b>")
                    
                    # Agregar liked_text sin traducción
                    if liked_text:
                        review_parts.append(f"👍 {liked_text}")
                    
                    # Agregar disliked_text sin traducción
                    if disliked_text:
                        review_parts.append(f"👎 {disliked_text}")
                    
                    # Si no hay texto de review, usar información de la estadía
                    if not review_parts:
                        # Construir descripción basada en la estadía (EN ALEMÁN)
                        stay_parts = []
                        
                        # Traducir tipo de viajero al alemán
                        traveler_type_de = translate_traveler_type(traveler_type)
                        if traveler_type_de:
                            stay_parts.append(f"<i>{traveler_type_de}</i>")
                        
                        # Procesar room_info (puede tener "1 rooms, 2 guests" por ejemplo)
                        if room_info:
                            room_info_de = translate_room_info(room_info)
                            stay_parts.append(f"<b>{room_info_de}</b>")
                        
                        # Usar rating como resumen si no hay otro texto
                        if stay_parts:
                            review_parts.append(' - '.join(stay_parts))
                        else:
                            # Si tampoco hay información de estadía, crear mensaje neutral basado en rating
                            review_parts.append(f"Bewertung: <b>{rating_booking}/10</b>")
                    
                    review_text = '<br/>'.join(review_parts) if review_parts else f"Bewertung: <b>{rating_booking}/10</b>"
                    
                    # Validación final: evitar reviews completamente vacías
                    if not review_text or (len(review_text) < 5 and not review_title and not liked_text and not disliked_text):
                        reviews_filtered_by_empty_content += 1
                        continue  # Saltar si el texto es muy corto y no hay componentes de review
                    
                    language = review.get('reviewLanguage', 'de')
                    review_id_booking = review.get('id', '')
                    
                    # Obtener código del apartamento desde URL de Booking
                    apartment_code = get_apartment_code_from_booking_url(booking_url)
                    if not apartment_code:
                        continue
                    
                    # Usar el código como apartment_id para Booking
                    apartment_id = f"booking_{apartment_code}"
                    unique_apartments.add(apartment_id)
                    
                    # Convertir fecha ISO a formato legible
                    date_display = convert_iso_to_display(review_date_iso)
                    
                    # Filtrar reviews antiguas
                    if not is_review_recent_iso(review_date_iso, max_days=MAX_REVIEW_AGE_DAYS):
                        reviews_filtered_by_date += 1
                        continue
                    
                    # Nombre del apartamento es el código
                    apartment_name = apartment_code
                    
                    # Inicializar estadísticas del apartamento
                    if apartment_id not in apartment_stats:
                        apartment_stats[apartment_id] = {'count': 0, 'name': apartment_name, 'source': 'Booking'}
                    
                    # Generar ID único
                    review_id = f"booking_{review_id_booking}_{apartment_code}"
                    
                    # Procesar rating (Booking usa escala de 10, mantener original)
                    if isinstance(rating_booking, (int, float)):
                        rating = int(rating_booking)  # Mantener escala de 10
                    else:
                        rating = 0
                    
                    # Información de estadía para Booking
                    stay_info = {
                        'check_in': review.get('checkInDate', ''),
                        'check_out': review.get('checkOutDate', ''),
                        'number_of_nights': review.get('numberOfNights', 0),
                        'room_info': review.get('roomInfo', ''),
                        'traveler_type': review.get('travelerType', '')
                    }
                    
                    processed_review = {
                        'id': review_id,
                        'text': review_text,
                        'rating': rating,
                        'max_rating': 10,  # Escala máxima de Booking
                        'platform_code': 'BK',  # Siglas para Booking
                        'date': date_display,
                        'reviewer_name': reviewer_name if reviewer_name else 'Anonym',
                        'reviewer_picture': reviewer_picture if reviewer_picture else '',
                        'reviewer_location': review.get('userLocation', ''),
                        'host_name': 'Urlaubsmagie',
                        'created_at': review_date_iso,
                        'language': language,
                        'apartment_id': apartment_id,
                        'apartment_name': apartment_name,
                        'apartment_url': booking_url,
                        'source': 'Booking',
                        'stay_info': stay_info  # Información de estadía solo para Booking
                    }
                    
                    processed_reviews.append(processed_review)
                    apartment_stats[apartment_id]['count'] += 1
        else:
            print(f"\n⚠️  Archivos Booking no encontrados")
        
        # Ordenar de forma aleatoria
        random.shuffle(processed_reviews)
        
        # Resumen final
        total_reviews = len(processed_reviews)
        total_apartments = len(unique_apartments)
        total_reviews_found = total_reviews_found_airbnb + total_reviews_found_booking
        
        print(f"\n📊 Resumen de carga:")
        print(f"   🏠 Total apartamentos únicos: {total_apartments}")
        print(f"   📄 Total comentarios encontrados:")
        print(f"      - Airbnb: {total_reviews_found_airbnb}")
        print(f"      - Booking: {total_reviews_found_booking}")
        print(f"      - Total: {total_reviews_found}")
        print(f"   🚭 Apartamentos excluidos: {reviews_filtered_by_apartment} comentarios")
        print(f"   📅 Filtrados por fecha (>{MAX_REVIEW_AGE_DAYS} días): {reviews_filtered_by_date} comentarios")
        print(f"   ⚠️  Filtrados por contenido vacío (Booking): {reviews_filtered_by_empty_content} comentarios")
        print(f"   ⭐ Comentarios finales (recientes, ≤{MAX_REVIEW_AGE_DAYS} días): {total_reviews}")
        print(f"\n🏠 Apartamentos con comentarios recientes:")
        for apt_id, info in apartment_stats.items():
            source_icon = "🅰️" if info['source'] == 'Airbnb' else "🅱️"
            print(f"   {source_icon} {info['name']}: {info['count']} comentarios")
        print()
        
        # Guardar en caché
        reviews_cache['neuen']['data'] = processed_reviews
        reviews_cache['neuen']['timestamp'] = datetime.now()
        print("✅ Reviews guardadas en caché\n")
        
        return processed_reviews
        
    except json.JSONDecodeError as e:
        print(f"❌ Error al leer JSON: {e}")
        return []
    except Exception as e:
        print(f"❌ Error crítico al cargar reviews: {e}")
        import traceback
        traceback.print_exc()
        return []

def load_historical_reviews():
    """Cargar todas las reviews históricas desde data/DataProblemListing (sin filtro de fecha)"""
    # Verificar si hay datos en caché válidos
    now = datetime.now()
    if reviews_cache['allem']['data'] is not None and reviews_cache['allem']['timestamp'] is not None:
        elapsed = (now - reviews_cache['allem']['timestamp']).total_seconds()
        if elapsed < CACHE_TTL:
            print(f"\n⚡ Usando reviews históricas en caché (edad: {int(elapsed)}s)")
            # Crear copia y mezclar aleatoriamente para mantener aleatoriedad
            cached_reviews = reviews_cache['allem']['data'].copy()
            random.shuffle(cached_reviews)
            return cached_reviews
    
    print("\n🔄 Cargando reviews históricas desde archivos...")
    processed_reviews = []
    total_airbnb_files = 0
    total_booking_files = 0
    total_reviews_airbnb = 0
    total_reviews_booking = 0
    
    try:
        if not os.path.exists(HISTORICAL_REVIEWS_DIR):
            print(f"⚠️  Directorio de reviews históricas no encontrado: {HISTORICAL_REVIEWS_DIR}")
            return []
        
        print(f"\n📁 Cargando desde: {HISTORICAL_REVIEWS_DIR}")
        
        # Listar todos los archivos JSON
        all_files = [f for f in os.listdir(HISTORICAL_REVIEWS_DIR) if f.endswith('.json')]
        
        for filename in all_files:
            file_path = os.path.join(HISTORICAL_REVIEWS_DIR, filename)
            
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    reviews_data = json.load(file)
                
                if not isinstance(reviews_data, list):
                    continue
                
                # Determinar si es Airbnb o Booking por el nombre del archivo
                is_airbnb = filename.startswith('Airbnb')
                is_booking = filename.startswith('Booking')
                
                if not is_airbnb and not is_booking:
                    continue
                
                # Extraer código del apartamento del nombre del archivo
                # Ejemplos: AirbnbB2.json -> B2, BookingF1.json -> F1
                apartment_code = filename.replace('Airbnb', '').replace('Booking', '').replace('.json', '')
                
                if is_airbnb:
                    total_airbnb_files += 1
                    # Procesar reviews de Airbnb
                    for review in reviews_data:
                        listing_url = review.get('listingUrl', '')
                        reviewer_name = review.get('reviewerName', 'Anonym')
                        reviewer_picture = review.get('reviewerProfilePicture', '')
                        review_date_iso = review.get('reviewDate', '')
                        review_text = review.get('reviewText', '')
                        language = review.get('language', 'de')
                        review_id_num = review.get('reviewId', 0)
                        rating = review.get('rating', 0)
                        
                        # Extraer apartment_id de la URL
                        apartment_id = extract_apartment_id_from_url(listing_url)
                        if not apartment_id:
                            continue
                        
                        # Filtrar apartamentos excluidos
                        if apartment_id in EXCLUDED_APARTMENT_IDS:
                            continue
                        
                        # Convertir fecha ISO a formato legible
                        date_display = convert_iso_to_display(review_date_iso)
                        
                        # Obtener nombre del apartamento (usar código del archivo como fallback)
                        apartment_name = get_apartment_name_from_url(listing_url)
                        if apartment_name.startswith('Wohnung'):
                            apartment_name = apartment_code
                        
                        # Generar ID único
                        review_id = f"airbnb_hist_{review_id_num}_{apartment_id}"
                        
                        # Procesar rating (Airbnb usa escala de 5)
                        if not isinstance(rating, int):
                            rating = int(rating) if rating else 0
                        
                        processed_review = {
                            'id': review_id,
                            'text': review_text,
                            'rating': rating,
                            'max_rating': 5,
                            'platform_code': 'AB',
                            'date': date_display,
                            'reviewer_name': reviewer_name,
                            'reviewer_picture': reviewer_picture,
                            'reviewer_location': '',
                            'host_name': 'Urlaubsmagie',
                            'created_at': review_date_iso,
                            'language': language,
                            'apartment_id': apartment_id,
                            'apartment_name': apartment_name,
                            'apartment_url': listing_url,
                            'source': 'Airbnb',
                            'stay_info': None
                        }
                        
                        processed_reviews.append(processed_review)
                        total_reviews_airbnb += 1
                
                elif is_booking:
                    total_booking_files += 1
                    # Procesar reviews de Booking
                    for review in reviews_data:
                        reviewer_name = review.get('userName', 'Anonym')
                        review_date_iso = review.get('reviewDate', '')
                        rating_booking = review.get('rating', 0)
                        
                        # Validar que haya rating válido
                        if not rating_booking or (isinstance(rating_booking, (int, float)) and rating_booking == 0):
                            continue
                        
                        # Construir texto de review de Booking
                        review_title = review.get('reviewTitle')
                        liked_text = review.get('likedText')
                        disliked_text = review.get('dislikedText')
                        traveler_type = review.get('travelerType', '')
                        
                        # Combinar textos de review (sin traducción para optimizar velocidad)
                        review_parts = []
                        
                        if review_title:
                            review_parts.append(f"<b>{review_title}</b>")
                        
                        if liked_text:
                            review_parts.append(f"👍 {liked_text}")
                        
                        if disliked_text:
                            review_parts.append(f"👎 {disliked_text}")
                        
                        # Si no hay texto, crear descripción basada en la estadía
                        if not review_parts:
                            traveler_type_de = translate_traveler_type(traveler_type)
                            if traveler_type_de:
                                review_parts.append(f"<i>{traveler_type_de}</i>")
                            else:
                                review_parts.append(f"Bewertung: <b>{rating_booking}/10</b>")
                        
                        review_text = '<br/>'.join(review_parts) if review_parts else f"Bewertung: <b>{rating_booking}/10</b>"
                        
                        # Validación final
                        if not review_text or len(review_text) < 5:
                            continue
                        
                        # Convertir fecha ISO a formato legible
                        date_display = convert_iso_to_display(review_date_iso)
                        
                        # Generar ID único
                        review_id = f"booking_hist_{apartment_code}_{review_date_iso}"
                        
                        # Procesar rating (Booking usa escala de 10)
                        if isinstance(rating_booking, (int, float)):
                            rating = int(rating_booking)
                        else:
                            rating = 0
                        
                        processed_review = {
                            'id': review_id,
                            'text': review_text,
                            'rating': rating,
                            'max_rating': 10,
                            'platform_code': 'BK',
                            'date': date_display,
                            'reviewer_name': reviewer_name if reviewer_name else 'Anonym',
                            'reviewer_picture': '',
                            'reviewer_location': '',
                            'host_name': 'Urlaubsmagie',
                            'created_at': review_date_iso,
                            'language': 'de',
                            'apartment_id': f"booking_{apartment_code}",
                            'apartment_name': apartment_code,
                            'apartment_url': '',
                            'source': 'Booking',
                            'stay_info': {
                                'number_of_nights': review.get('numberOfNights', 0),
                                'traveler_type': traveler_type
                            }
                        }
                        
                        processed_reviews.append(processed_review)
                        total_reviews_booking += 1
            
            except Exception as e:
                print(f"   [!] Error procesando {filename}: {e}")
                continue
        
        # Ordenar de forma aleatoria
        random.shuffle(processed_reviews)
        
        # Resumen
        print(f"\n📊 Resumen de reviews históricas:")
        print(f"   📁 Archivos Airbnb: {total_airbnb_files}")
        print(f"   📁 Archivos Booking: {total_booking_files}")
        print(f"   ⭐ Reviews Airbnb: {total_reviews_airbnb}")
        print(f"   ⭐ Reviews Booking: {total_reviews_booking}")
        print(f"   📋 Total reviews históricas: {len(processed_reviews)}")
        
        # Guardar en caché
        reviews_cache['allem']['data'] = processed_reviews
        reviews_cache['allem']['timestamp'] = datetime.now()
        print("✅ Reviews históricas guardadas en caché\n")
        
        return processed_reviews
        
    except Exception as e:
        print(f"❌ Error crítico cargando reviews históricas: {e}")
        import traceback
        traceback.print_exc()
        return []

def load_reviews_for_slideshow():
    """Cargar reviews CON traducción automática para el slideshow (TV)"""
    # Verificar si hay datos en caché válidos para slideshow
    cache_key = 'slideshow'
    if cache_key not in reviews_cache:
        reviews_cache[cache_key] = {'data': None, 'timestamp': None}
    
    now = datetime.now()
    if reviews_cache[cache_key]['data'] is not None and reviews_cache[cache_key]['timestamp'] is not None:
        elapsed = (now - reviews_cache[cache_key]['timestamp']).total_seconds()
        if elapsed < CACHE_TTL:
            print(f"\n⚡ Usando reviews slideshow en caché (edad: {int(elapsed)}s)")
            # Crear copia y mezclar aleatoriamente
            cached_reviews = reviews_cache[cache_key]['data'].copy()
            random.shuffle(cached_reviews)
            return cached_reviews
    
    print("\n🔄 Cargando reviews para slideshow (con traducción)...")
    
    # Obtener reviews sin traducir del caché neuen
    base_reviews = load_reviews()
    
    # Traducir solo los textos de Booking que no estén en alemán
    translated_reviews = []
    for review in base_reviews:
        review_copy = review.copy()
        
        # Si es de Booking y tiene HTML tags, traducir
        if review_copy['source'] == 'Booking' and '<b>' in review_copy['text']:
            try:
                # Extraer y traducir cada parte
                text_parts = []
                
                # Parsear el HTML para obtener los textos
                import html
                from html.parser import HTMLParser
                
                class TextExtractor(HTMLParser):
                    def __init__(self):
                        super().__init__()
                        self.texts = []
                        self.in_bold = False
                    
                    def handle_starttag(self, tag, attrs):
                        if tag == 'b':
                            self.in_bold = True
                    
                    def handle_endtag(self, tag):
                        if tag == 'b':
                            self.in_bold = False
                    
                    def handle_data(self, data):
                        clean_data = data.strip()
                        if clean_data and clean_data not in ['👍', '👎']:
                            self.texts.append((clean_data, self.in_bold))
                
                parser = TextExtractor()
                parser.feed(review_copy['text'])
                
                # Traducir cada texto
                translated_parts = []
                for text, is_bold in parser.texts:
                    translated, _, _ = translate_to_german(text)
                    if is_bold:
                        translated_parts.append(f"<b>{translated}</b>")
                    else:
                        # Detectar si es liked o disliked por el emoji
                        if '👍' in review_copy['text'] and text in review_copy['text'].split('👍')[1]:
                            translated_parts.append(f"👍 {translated}")
                        elif '👎' in review_copy['text'] and text in review_copy['text'].split('👎')[1]:
                            translated_parts.append(f"👎 {translated}")
                        else:
                            translated_parts.append(translated)
                
                review_copy['text'] = '<br/>'.join(translated_parts)
                
            except Exception as e:
                print(f"⚠️  Error traduciendo review {review_copy['id']}: {e}")
                # Mantener texto original si falla
        
        translated_reviews.append(review_copy)
    
    # Mezclar aleatoriamente
    random.shuffle(translated_reviews)
    
    # Guardar en caché
    reviews_cache[cache_key]['data'] = translated_reviews
    reviews_cache[cache_key]['timestamp'] = datetime.now()
    print(f"✅ {len(translated_reviews)} reviews traducidas y guardadas en caché slideshow\n")
    
    return translated_reviews

def get_apartment_info():
    """Información del apartamento basada en el ID"""
    config = load_apartment_config()
    general_info = config.get('general_info', {})
    apartments = config.get('apartments', {})
    
    # Crear lista de códigos de apartamentos
    apartment_codes = [apt['code'] for apt in apartments.values()]
    
    return {
        'id': '1086254381830942280',
        'name': f'Apartments {", ".join(apartment_codes)}',
        'title': general_info.get('title', 'Ferienwohnungen - Sächsische Schweiz'),
        'location': general_info.get('location', 'Sebnitz, Sächsische Schweiz'),
        'description': f'Gemütliche Ferienwohnungen mit verschiedenen Ausstattungen: {", ".join(apartment_codes)}',
        'host': general_info.get('host', 'Urlaubsmagie')
    }

def get_statistics(reviews):
    """Calcular estadísticas de las reseñas"""
    if not reviews:
        return {
            'total_reviews': 0,
            'average_rating': 0,
            'rating_distribution': {1: 0, 2: 0, 3: 0, 4: 0, 5: 0},
            'five_star_percentage': 0
        }
    
    total_reviews = len(reviews)
    # Filtrar ratings válidos (no None, no 0, y que sean números)
    ratings = []
    for r in reviews:
        rating = r.get('rating')
        if rating is not None and isinstance(rating, (int, float)) and rating > 0:
            ratings.append(int(rating))
    
    avg_rating = sum(ratings) / len(ratings) if ratings else 0
    
    # Contar distribución de estrellas
    rating_distribution = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for rating in ratings:
        if rating in rating_distribution:
            rating_distribution[rating] += 1
    
    return {
        'total_reviews': total_reviews,
        'average_rating': round(avg_rating, 1),
        'rating_distribution': rating_distribution,
        'five_star_percentage': round((rating_distribution[5] / len(ratings)) * 100) if ratings else 0
    }

@app.route('/robots.txt')
def robots_txt():
    """Block all search engine crawlers - site is for internal company use only"""
    robots_content = """# Robots.txt - Urlaubsmagie Bewertungsportal
# This site is for internal company use only
# All search engine crawlers are blocked

User-agent: *
Disallow: /

# Block all known search engines explicitly
User-agent: Googlebot
Disallow: /

User-agent: Bingbot
Disallow: /

User-agent: Yahoo! Slurp
Disallow: /

User-agent: DuckDuckBot
Disallow: /

User-agent: Baiduspider
Disallow: /

User-agent: YandexBot
Disallow: /

User-agent: facebot
Disallow: /

User-agent: ia_archiver
Disallow: /
"""
    return robots_content, 200, {'Content-Type': 'text/plain'}

@app.route('/')
def homepage():
    """Página principal - Menú de navegación"""
    return render_template('homepage.html')

@app.route('/home')
def home_alias():
    """Alias para la página principal"""
    return render_template('homepage.html')

@app.route('/reviews')
def reviews_view():
    """Vista de reseñas con filtros y búsqueda - Soporta categorías neuen/allem"""
    # Obtener categoría desde query parameter (default: neuen)
    category = request.args.get('category', 'neuen').lower()
    
    # Cargar reviews según la categoría
    if category == 'allem':
        reviews = load_historical_reviews()
        category_label = 'Allem'
    else:
        reviews = load_reviews()
        category_label = 'Neuen'
    
    stats = get_statistics(reviews)
    
    return render_template('index.html',
                         reviews=reviews,
                         stats=stats,
                         total_reviews=len(reviews),
                         category=category,
                         category_label=category_label)

def calculate_apartment_statistics(reviews):
    """Calcular estadísticas agrupadas por apartamento"""
    apartment_data = {}
    
    for review in reviews:
        apt_id = review.get('apartment_id', 'unknown')
        apt_name = review.get('apartment_name', 'Unknown')
        
        if apt_id not in apartment_data:
            apartment_data[apt_id] = {
                'name': apt_name,
                'reviews': [],
                'airbnb_reviews': [],
                'booking_reviews': []
            }
        
        apartment_data[apt_id]['reviews'].append(review)
        
        if review.get('source') == 'Airbnb':
            apartment_data[apt_id]['airbnb_reviews'].append(review)
        else:
            apartment_data[apt_id]['booking_reviews'].append(review)
    
    # Calcular estadísticas por cada apartamento
    result = {}
    for apt_id, data in apartment_data.items():
        reviews_list = data['reviews']
        airbnb_reviews = data['airbnb_reviews']
        booking_reviews = data['booking_reviews']
        
        # Total de reviews
        total = len(reviews_list)
        
        # Rating promedio
        ratings = [r.get('rating', 0) for r in reviews_list if r.get('rating', 0) > 0]
        avg_rating = sum(ratings) / len(ratings) if ratings else 0
        
        # Distribución por estrellas
        distribution = {}
        for star in range(1, 6):
            distribution[str(star)] = sum(1 for r in reviews_list if r.get('rating') == star)
        
        # Promedios por plataforma
        airbnb_ratings = [r.get('rating', 0) for r in airbnb_reviews if r.get('rating', 0) > 0]
        booking_ratings = [r.get('rating', 0) for r in booking_reviews if r.get('rating', 0) > 0]
        
        airbnb_avg = sum(airbnb_ratings) / len(airbnb_ratings) if airbnb_ratings else 0
        booking_avg = sum(booking_ratings) / len(booking_ratings) if booking_ratings else 0
        
        # Preparar datos serializables para JavaScript
        all_reviews_data = []
        for r in reviews_list:
            all_reviews_data.append({
                'rating': r.get('rating', 0),
                'source': r.get('source', ''),
                'created_at': r.get('created_at', '')
            })
        
        result[apt_id] = {
            'name': data['name'],
            'total_reviews': total,
            'average_rating': avg_rating,
            'rating_distribution': distribution,
            'airbnb_count': len(airbnb_reviews),
            'booking_count': len(booking_reviews),
            'airbnb_avg': airbnb_avg,
            'booking_avg': booking_avg,
            'all_reviews': all_reviews_data  # Para filtros dinámicos en JS
        }
    
    # Ordenar por rating promedio (mejor primero)
    result = dict(sorted(result.items(), key=lambda x: x[1]['average_rating'], reverse=True))
    
    return result

@app.route('/reviews/digital-team')
def digital_team_view():
    """Vista especial para el equipo digital"""
    # Cargar todos los archivos DTAirbnb y DTBooking
    dt_reviews = load_digital_team_reviews()
    
    # Agregar fecha exacta formateada a cada review
    for review in dt_reviews:
        if review.get('created_at'):
            review['exact_date_formatted'] = format_exact_date(review['created_at'])
    
    stats = get_statistics(dt_reviews)
    
    # Obtener años y meses disponibles para filtros
    available_dates = get_available_dates_from_reviews(dt_reviews)
    
    # Calcular estadísticas por apartamento
    apartment_stats = calculate_apartment_statistics(dt_reviews)
    
    return render_template('digital_team.html',
                         reviews=dt_reviews,
                         stats=stats,
                         total_reviews=len(dt_reviews),
                         available_years=available_dates['years'],
                         available_months=available_dates['months'],
                         apartment_stats=apartment_stats)

def format_exact_date(iso_date_str):
    """Formatear fecha ISO a formato legible en alemán
    Ejemplo: 2025-12-28T15:00:55Z -> 28.12.2025 um 15:00 Uhr
    """
    try:
        dt = datetime.fromisoformat(iso_date_str.replace('Z', '+00:00'))
        return dt.strftime('%d.%m.%Y um %H:%M Uhr')
    except:
        return iso_date_str

def load_digital_team_reviews():
    """Cargar reviews de archivos DTAirbnb, DTBooking, AirbnbDataDT y BookingDataDT"""
    dt_folder = r"C:\n8n_Docker\Files"
    processed_reviews = []

    # Patrón para encontrar archivos DT: DTAirbnb_DD_MM_YY_HHMMSS.json o DTBooking_DD_MM_YY_HHMMSS.json
    import glob

    dt_airbnb_files = glob.glob(os.path.join(dt_folder, "DTAirbnb_*.json"))
    dt_booking_files = glob.glob(os.path.join(dt_folder, "DTBooking_*.json"))

    # Agregar archivos de historial completo
    airbnb_data_dt = os.path.join(dt_folder, "AirbnbDataDT.json")
    booking_data_dt = os.path.join(dt_folder, "BookingDataDT.json")

    if os.path.exists(airbnb_data_dt):
        dt_airbnb_files.append(airbnb_data_dt)
    if os.path.exists(booking_data_dt):
        dt_booking_files.append(booking_data_dt)

    print(f"\n[DT] Cargando Digital Team reviews...")
    print(f"   [AB] Archivos Airbnb encontrados: {len(dt_airbnb_files)}")
    print(f"   [BK] Archivos Booking encontrados: {len(dt_booking_files)}")

    # Usar conjuntos para evitar duplicados entre todos los archivos
    seen_airbnb_review_ids = set()
    seen_booking_review_ids = set()

    # Procesar archivos DTAirbnb
    for file_path in dt_airbnb_files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                airbnb_data = json.load(f)

            # Extraer fecha del nombre del archivo: DTAirbnb_06_01_26_113415.json
            filename = os.path.basename(file_path)
            file_created_at = parse_dt_filename(filename)

            print(f"      > Procesando: {filename} ({len(airbnb_data)} reviews)")

            for review in airbnb_data:
                review_id = review.get('reviewId', '')

                # Evitar duplicados (especialmente importante con archivos de historial completo)
                if review_id and review_id in seen_airbnb_review_ids:
                    continue
                if review_id:
                    seen_airbnb_review_ids.add(review_id)

                listing_url = review.get('listingUrl', '')
                apartment_id = extract_apartment_id_from_url(listing_url)

                # Filtrar apartamentos excluidos
                if apartment_id in EXCLUDED_APARTMENT_IDS:
                    continue

                apartment_name = get_apartment_name_from_url(listing_url) if apartment_id else 'Unknown'

                review_date_iso = review.get('reviewDate', '')
                date_display = convert_iso_to_display(review_date_iso)

                processed_review = {
                    'id': f"dt_airbnb_{review_id}",
                    'text': review.get('reviewText', ''),
                    'rating': review.get('rating', 0),
                    'max_rating': 5,
                    'platform_code': 'AB',
                    'date': date_display,
                    'reviewer_name': review.get('reviewerName', 'Anonym'),
                    'reviewer_picture': review.get('reviewerProfilePicture', ''),
                    'reviewer_location': '',
                    'host_name': 'Urlaubsmagie',
                    'created_at': review_date_iso,
                    'language': review.get('language', 'de'),
                    'apartment_id': apartment_id or 'unknown',
                    'apartment_name': apartment_name,
                    'apartment_url': listing_url,
                    'source': 'Airbnb',
                    'file_created_at': file_created_at,
                    'stay_info': None
                }
                processed_reviews.append(processed_review)
        except Exception as e:
            print(f"   [!] Error procesando {filename}: {e}")
    
    # Procesar archivos DTBooking
    for file_path in dt_booking_files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                booking_data = json.load(f)

            filename = os.path.basename(file_path)
            file_created_at = parse_dt_filename(filename)

            print(f"      > Procesando: {filename} ({len(booking_data)} reviews)")

            for review in booking_data:
                # Crear un ID único para evitar duplicados en Booking (no tienen reviewId como Airbnb)
                # Usamos: userName + reviewDate + rating como identificador compuesto
                review_unique_key = f"{review.get('userName', '')}_{review.get('reviewDate', '')}_{review.get('rating', '')}"

                # Evitar duplicados
                if review_unique_key in seen_booking_review_ids:
                    continue
                seen_booking_review_ids.add(review_unique_key)

                booking_url = review.get('startUrl', '')
                apartment_code = get_apartment_code_from_booking_url(booking_url)

                # Limpiar URL removiendo el código entre paréntesis al final
                booking_url_clean = re.sub(r'\s*\([^)]+\)\s*$', '', booking_url)

                review_date_iso = review.get('reviewDate', '')
                date_display = convert_iso_to_display(review_date_iso)
                
                rating_booking = review.get('rating', 0)
                
                # Construir texto de review de Booking
                review_parts = []
                review_title = review.get('reviewTitle')
                liked_text = review.get('likedText')
                disliked_text = review.get('dislikedText')
                
                if review_title:
                    review_parts.append(f"<b>{review_title}</b>")
                if liked_text:
                    review_parts.append(f"👍 {liked_text}")
                if disliked_text:
                    review_parts.append(f"👎 {disliked_text}")
                
                review_text = '<br/>'.join(review_parts) if review_parts else f"Bewertung: <b>{rating_booking}/10</b>"
                
                # Normalizar rating de Booking (10) a escala de Airbnb (5)
                # 8-10 = 5 estrellas, 6-7 = 4 estrellas, 4-5 = 3 estrellas, 2-3 = 2 estrellas, 0-1 = 1 estrella
                if rating_booking >= 8:
                    rating_normalized = 5
                elif rating_booking >= 6:
                    rating_normalized = 4
                elif rating_booking >= 4:
                    rating_normalized = 3
                elif rating_booking >= 2:
                    rating_normalized = 2
                elif rating_booking >= 1:
                    rating_normalized = 1
                else:
                    rating_normalized = 0
                
                processed_review = {
                    'id': f"dt_booking_{hash(review_unique_key)}",
                    'text': review_text,
                    'rating': rating_normalized,  # Normalizado a escala de 5
                    'rating_original': int(rating_booking) if rating_booking else 0,  # Original de 10
                    'max_rating': 5,  # Ahora también usa escala de 5
                    'platform_code': 'BK',
                    'date': date_display,
                    'reviewer_name': review.get('userName', 'Anonym'),
                    'reviewer_picture': review.get('userAvatar', ''),
                    'reviewer_location': review.get('userLocation', ''),
                    'host_name': 'Urlaubsmagie',
                    'created_at': review_date_iso,
                    'language': review.get('reviewLanguage', 'de'),
                    'apartment_id': f"booking_{apartment_code}" if apartment_code else 'unknown',
                    'apartment_name': apartment_code or 'Unknown',
                    'apartment_url': booking_url_clean,  # URL limpia sin código
                    'source': 'Booking',
                    'file_created_at': file_created_at,
                    'stay_info': {
                        'number_of_nights': review.get('numberOfNights', 0),
                        'traveler_type': review.get('travelerType', '')
                    }
                }
                processed_reviews.append(processed_review)
        except Exception as e:
            print(f"   [!] Error procesando {filename}: {e}")

    # Contar reviews por fuente
    airbnb_count = sum(1 for r in processed_reviews if r['source'] == 'Airbnb')
    booking_count = sum(1 for r in processed_reviews if r['source'] == 'Booking')

    print(f"[OK] Total Digital Team reviews: {len(processed_reviews)}")
    print(f"   [AB] Airbnb: {airbnb_count} reviews")
    print(f"   [BK] Booking: {booking_count} reviews\n")
    return processed_reviews

def parse_dt_filename(filename):
    """Parsear nombre de archivo DT para extraer fecha y hora
    Ejemplo: DTAirbnb_06_01_26_113415.json -> 2026-01-06T11:34:15Z
    Para archivos de historial completo (AirbnbDataDT.json, BookingDataDT.json) usa la fecha actual
    """
    try:
        # Si es un archivo de historial completo (sin timestamp en el nombre)
        if filename in ['AirbnbDataDT.json', 'BookingDataDT.json']:
            return 'history_full'  # Marcador especial para archivos de historial

        # Remover extensión y prefijo
        parts = filename.replace('.json', '').split('_')
        if len(parts) >= 5:
            day = parts[1]
            month = parts[2]
            year = f"20{parts[3]}"  # Convertir 26 -> 2026
            time_str = parts[4]  # 113415

            hour = time_str[:2]
            minute = time_str[2:4]
            second = time_str[4:6]

            return f"{year}-{month}-{day}T{hour}:{minute}:{second}Z"
    except:
        return datetime.now().isoformat()
    return datetime.now().isoformat()

def get_available_dates_from_reviews(reviews):
    """Extraer años y meses disponibles de las reviews"""
    years = set()
    months = set()
    
    for review in reviews:
        try:
            created_at = review.get('created_at', '')
            if created_at:
                dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                years.add(dt.year)
                months.add(dt.month)
        except:
            continue
    
    return {
        'years': sorted(list(years), reverse=True),
        'months': sorted(list(months))
    }

@app.route('/api/digital-team/export-excel')
def export_digital_team_excel():
    """Exportar reviews filtradas de Digital Team a Excel"""
    try:
        # Obtener parámetros de filtro
        search_term = request.args.get('search', '').lower()
        platform_filter = request.args.get('platform', '')
        rating_filter = request.args.get('rating', '')
        month_filter = request.args.get('month', '')
        year_filter = request.args.get('year', '')
        sort_order = request.args.get('sort', 'newest')
        
        # Cargar todas las reviews
        dt_reviews = load_digital_team_reviews()
        
        # Aplicar filtros
        filtered_reviews = []
        for review in dt_reviews:
            # Filtro de búsqueda
            if search_term and search_term not in review.get('text', '').lower():
                continue
            
            # Filtro de plataforma
            if platform_filter and review.get('source', '') != platform_filter:
                continue
            
            # Filtro de rating (ya normalizado en el backend)
            if rating_filter:
                review_rating = review.get('rating', 0)
                if review_rating != int(rating_filter):
                    continue
            
            # Filtro de mes y año
            if month_filter or year_filter:
                try:
                    created_at = review.get('created_at', '')
                    if created_at:
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        if month_filter and dt.month != int(month_filter):
                            continue
                        if year_filter and dt.year != int(year_filter):
                            continue
                except:
                    continue
            
            filtered_reviews.append(review)
        
        # Ordenar
        if sort_order == 'newest':
            filtered_reviews.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        elif sort_order == 'oldest':
            filtered_reviews.sort(key=lambda x: x.get('created_at', ''))
        elif sort_order == 'rating-high':
            filtered_reviews.sort(key=lambda x: x.get('rating', 0), reverse=True)
        elif sort_order == 'rating-low':
            filtered_reviews.sort(key=lambda x: x.get('rating', 0))
        
        # Crear archivo Excel
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Digital Team Reviews"
        
        # Encabezados
        headers = ['Plattform', 'Name', 'Standort', 'Bewertung', 'Datum', 'Exakte Datum', 'Apartment', 'Bewertungstext']
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Agregar datos
        for review in filtered_reviews:
            # Limpiar texto HTML
            import re
            text = review.get('text', '')
            text = re.sub(r'<br/>', ' ', text)
            text = re.sub(r'<[^>]+>', '', text)
            text = text.replace('👍', '[Me gusta]').replace('👎', '[No me gusta]')
            
            # Mostrar rating según la plataforma
            rating = review.get('rating', 0)
            rating_original = review.get('rating_original', 0)
            
            if rating_original and review.get('source') == 'Booking':
                rating = f"{rating}/5 ({rating_original}/10)"
            else:
                rating = f"{rating}/5"
            
            # Fecha exacta formateada
            exact_date = format_exact_date(review.get('created_at', '')) if review.get('created_at') else ''
            
            row = [
                review.get('source', ''),
                review.get('reviewer_name', ''),
                review.get('reviewer_location', ''),
                rating,
                review.get('date', ''),
                exact_date,
                review.get('apartment_name', ''),
                text
            ]
            ws.append(row)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12  # Plattform
        ws.column_dimensions['B'].width = 20  # Name
        ws.column_dimensions['C'].width = 15  # Standort
        ws.column_dimensions['D'].width = 15  # Bewertung
        ws.column_dimensions['E'].width = 20  # Datum
        ws.column_dimensions['F'].width = 25  # Exakte Datum
        ws.column_dimensions['G'].width = 15  # Apartment
        ws.column_dimensions['H'].width = 80  # Bewertungstext
        
        # Wrap text en columna de texto
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[7].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo con filtros aplicados
        filename_parts = ['Digital_Team_Reviews']
        if platform_filter:
            filename_parts.append(platform_filter)
        if rating_filter:
            filename_parts.append(f"{rating_filter}Sterne")
        if month_filter:
            month_names = ['Jan', 'Feb', 'Maerz', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']
            filename_parts.append(month_names[int(month_filter)-1])
        if year_filter:
            filename_parts.append(year_filter)
        
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Error exportando a Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/digital-team/export-apartments-excel')
def export_apartments_excel():
    """Exportar estadísticas de apartamentos a Excel"""
    try:
        # Obtener parámetros de filtro
        platform_filter = request.args.get('platform', '')
        month_filter = request.args.get('month', '')
        year_filter = request.args.get('year', '')
        
        # Cargar todas las reviews
        dt_reviews = load_digital_team_reviews()
        
        # Calcular estadísticas de apartamentos
        apartment_stats = calculate_apartment_statistics(dt_reviews)
        
        # Aplicar filtros y recalcular estadísticas
        filtered_apartment_stats = {}
        
        for apt_id, apt_data in apartment_stats.items():
            all_reviews = apt_data.get('all_reviews', [])
            
            # Filtrar reviews
            filtered_reviews = []
            for review in all_reviews:
                # Filtro de plataforma
                if platform_filter and review.get('source', '') != platform_filter:
                    continue
                
                # Filtro de mes y año
                if month_filter or year_filter:
                    try:
                        created_at = review.get('created_at', '')
                        if created_at:
                            dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                            if month_filter and dt.month != int(month_filter):
                                continue
                            if year_filter and dt.year != int(year_filter):
                                continue
                    except:
                        continue
                
                filtered_reviews.append(review)
            
            # Solo incluir apartamentos con reviews después del filtro
            if filtered_reviews:
                # Recalcular estadísticas
                total_reviews = len(filtered_reviews)
                total_rating = sum(r.get('rating', 0) for r in filtered_reviews)
                avg_rating = total_rating / total_reviews if total_reviews > 0 else 0
                
                # Distribución de estrellas
                distribution = {'1': 0, '2': 0, '3': 0, '4': 0, '5': 0}
                for r in filtered_reviews:
                    rating = str(int(r.get('rating', 0)))
                    if rating in distribution:
                        distribution[rating] += 1
                
                # Por plataforma
                airbnb_reviews = [r for r in filtered_reviews if r.get('source') == 'Airbnb']
                booking_reviews = [r for r in filtered_reviews if r.get('source') == 'Booking']
                
                airbnb_count = len(airbnb_reviews)
                booking_count = len(booking_reviews)
                airbnb_avg = sum(r.get('rating', 0) for r in airbnb_reviews) / airbnb_count if airbnb_count > 0 else 0
                booking_avg = sum(r.get('rating', 0) for r in booking_reviews) / booking_count if booking_count > 0 else 0
                
                filtered_apartment_stats[apt_id] = {
                    'name': apt_data.get('name', 'Unknown'),
                    'total_reviews': total_reviews,
                    'average_rating': avg_rating,
                    'distribution': distribution,
                    'airbnb_count': airbnb_count,
                    'booking_count': booking_count,
                    'airbnb_avg': airbnb_avg,
                    'booking_avg': booking_avg,
                    'star_5': distribution.get('5', 0),
                    'star_4': distribution.get('4', 0),
                    'star_3': distribution.get('3', 0),
                    'star_2': distribution.get('2', 0),
                    'star_1': distribution.get('1', 0)
                }
        
        # Crear archivo Excel
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Wohnungen Statistiken"
        
        # Encabezados
        headers = [
            'Wohnung', 
            'Gesamt Bewertungen', 
            'Durchschnitt',
            '5 Sterne',
            '4 Sterne',
            '3 Sterne',
            '2 Sterne',
            '1 Stern',
            'Airbnb Anzahl',
            'Airbnb Durchschnitt',
            'Booking Anzahl',
            'Booking Durchschnitt'
        ]
        
        # Añadir información de filtros si existen
        if month_filter or year_filter or platform_filter:
            filter_info = []
            if platform_filter:
                filter_info.append(f"Plattform: {platform_filter}")
            if month_filter:
                month_names = ['Januar', 'Februar', 'März', 'April', 'Mai', 'Juni', 'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']
                filter_info.append(f"Monat: {month_names[int(month_filter)-1]}")
            if year_filter:
                filter_info.append(f"Jahr: {year_filter}")
            
            ws.append(['Filter: ' + ', '.join(filter_info)])
            ws.merge_cells('A1:L1')
            ws['A1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.append([])  # Fila vacía
        
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        header_row = 3 if (month_filter or year_filter or platform_filter) else 1
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Agregar datos de apartamentos
        for apt_id, apt_data in sorted(filtered_apartment_stats.items(), key=lambda x: x[1]['name']):
            row = [
                apt_data['name'],
                apt_data['total_reviews'],
                round(apt_data['average_rating'], 2),
                apt_data['star_5'],
                apt_data['star_4'],
                apt_data['star_3'],
                apt_data['star_2'],
                apt_data['star_1'],
                apt_data['airbnb_count'],
                round(apt_data['airbnb_avg'], 2) if apt_data['airbnb_avg'] > 0 else 'N/A',
                apt_data['booking_count'],
                round(apt_data['booking_avg'], 2) if apt_data['booking_avg'] > 0 else 'N/A'
            ]
            ws.append(row)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20  # Wohnung
        ws.column_dimensions['B'].width = 18  # Gesamt Bewertungen
        ws.column_dimensions['C'].width = 15  # Durchschnitt
        ws.column_dimensions['D'].width = 12  # 5 Sterne
        ws.column_dimensions['E'].width = 12  # 4 Sterne
        ws.column_dimensions['F'].width = 12  # 3 Sterne
        ws.column_dimensions['G'].width = 12  # 2 Sterne
        ws.column_dimensions['H'].width = 12  # 1 Stern
        ws.column_dimensions['I'].width = 15  # Airbnb Anzahl
        ws.column_dimensions['J'].width = 20  # Airbnb Durchschnitt
        ws.column_dimensions['K'].width = 15  # Booking Anzahl
        ws.column_dimensions['L'].width = 20  # Booking Durchschnitt
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo con filtros aplicados
        filename_parts = ['Wohnungen_Statistiken']
        if platform_filter:
            filename_parts.append(platform_filter)
        if month_filter:
            month_names = ['Jan', 'Feb', 'Maerz', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']
            filename_parts.append(month_names[int(month_filter)-1])
        if year_filter:
            filename_parts.append(year_filter)
        
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Error exportando apartamentos a Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/digital-team/export-general-stats-excel')
def export_general_stats_excel():
    """Exportar estadísticas generales de Urlaubsmagie a Excel"""
    try:
        # Obtener parámetros de filtro
        platform_filter = request.args.get('platform', '')
        month_filter = request.args.get('month', '')
        year_filter = request.args.get('year', '')
        
        # Cargar todas las reviews
        dt_reviews = load_digital_team_reviews()
        
        # Aplicar filtros
        filtered_reviews = []
        for review in dt_reviews:
            # Filtro de plataforma
            if platform_filter and review.get('source', '') != platform_filter:
                continue
            
            # Filtro de mes y año
            if month_filter or year_filter:
                try:
                    created_at = review.get('created_at', '')
                    if created_at:
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        if month_filter and dt.month != int(month_filter):
                            continue
                        if year_filter and dt.year != int(year_filter):
                            continue
                except:
                    continue
            
            filtered_reviews.append(review)
        
        # Calcular estadísticas generales
        total_reviews = len(filtered_reviews)
        
        if total_reviews == 0:
            # Si no hay datos, crear Excel vacío con mensaje
            import io
            from openpyxl import Workbook
            from flask import send_file
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Gesamtstatistiken"
            ws.append(['Keine Daten verfügbar für die ausgewählten Filter'])
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='Urlaubsmagie_Gesamtstatistiken_Leer.xlsx'
            )
        
        # Calcular estadísticas
        total_rating = sum(r.get('rating', 0) for r in filtered_reviews)
        avg_rating = total_rating / total_reviews
        
        # Distribución de estrellas
        distribution = {'1': 0, '2': 0, '3': 0, '4': 0, '5': 0}
        for r in filtered_reviews:
            rating = str(int(r.get('rating', 0)))
            if rating in distribution:
                distribution[rating] += 1
        
        # Porcentaje de 5 estrellas
        five_star_count = distribution.get('5', 0)
        five_star_percent = (five_star_count / total_reviews * 100) if total_reviews > 0 else 0
        
        # Por plataforma
        airbnb_reviews = [r for r in filtered_reviews if r.get('source') == 'Airbnb']
        booking_reviews = [r for r in filtered_reviews if r.get('source') == 'Booking']
        
        airbnb_count = len(airbnb_reviews)
        booking_count = len(booking_reviews)
        airbnb_avg = sum(r.get('rating', 0) for r in airbnb_reviews) / airbnb_count if airbnb_count > 0 else 0
        booking_avg = sum(r.get('rating', 0) for r in booking_reviews) / booking_count if booking_count > 0 else 0
        
        # Crear archivo Excel
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Gesamtstatistiken"
        
        # Título principal
        ws.append(['Urlaubsmagie - Gesamtstatistiken'])
        ws.merge_cells('A1:B1')
        ws['A1'].font = Font(bold=True, size=16, color="667EEA")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.append([])  # Fila vacía
        
        # Información de filtros si existen
        if month_filter or year_filter or platform_filter:
            filter_info = []
            if platform_filter:
                filter_info.append(f"Plattform: {platform_filter}")
            if month_filter:
                month_names = ['Januar', 'Februar', 'März', 'April', 'Mai', 'Juni', 'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']
                filter_info.append(f"Monat: {month_names[int(month_filter)-1]}")
            if year_filter:
                filter_info.append(f"Jahr: {year_filter}")
            
            ws.append(['Filter: ' + ', '.join(filter_info)])
            ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
            ws[f'A{ws.max_row}'].font = Font(bold=True)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='center')
            ws.append([])  # Fila vacía
        
        # Sección: Totales principales
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws.append(['Metrik', 'Wert'])
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.append(['Gesamtbewertungen', total_reviews])
        ws.append(['Durchschnittsbewertung', round(avg_rating, 2)])
        ws.append(['5-Sterne-Bewertungen (%)', round(five_star_percent, 1)])
        ws.append([])  # Fila vacía
        
        # Sección: Distribución por estrellas
        ws.append(['Verteilung nach Sternen'])
        ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='center')
        
        ws.append(['Sterne', 'Anzahl'])
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for star in range(5, 0, -1):
            count = distribution.get(str(star), 0)
            percentage = (count / total_reviews * 100) if total_reviews > 0 else 0
            ws.append([f'{star} Sterne', f'{count} ({percentage:.1f}%)'])
        
        ws.append([])  # Fila vacía
        
        # Sección: Por plataforma
        ws.append(['Nach Plattform'])
        ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='center')
        
        ws.append(['Plattform', 'Details'])
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.append(['Airbnb', f'{airbnb_count} Bewertungen (Durchschnitt: {round(airbnb_avg, 2) if airbnb_avg > 0 else "N/A"})'])
        ws.append(['Booking', f'{booking_count} Bewertungen (Durchschnitt: {round(booking_avg, 2) if booking_avg > 0 else "N/A"})'])
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo con filtros aplicados
        filename_parts = ['Urlaubsmagie_Gesamtstatistiken']
        if platform_filter:
            filename_parts.append(platform_filter)
        if month_filter:
            month_names = ['Jan', 'Feb', 'Maerz', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']
            filename_parts.append(month_names[int(month_filter)-1])
        if year_filter:
            filename_parts.append(year_filter)
        
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Error exportando estadísticas generales a Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/slideshow')
def slideshow():
    """Presentación automática de reseñas para TV"""
    reviews = load_reviews()

    # Agregar fecha exacta formateada a cada review
    for review in reviews:
        if review.get('created_at'):
            review['exact_date_formatted'] = format_exact_date(review['created_at'])

    apartment_info = get_apartment_info()
    stats = get_statistics(reviews)

    return render_template('slideshow.html',
                         reviews=reviews,
                         apartment_info=apartment_info,
                         stats=stats,
                         total_reviews=len(reviews))

@app.route('/api/reviews')
def api_reviews():
    """API endpoint para obtener todas las reseñas"""
    reviews = load_reviews()
    return jsonify(reviews)

@app.route('/api/stats')
def api_stats():
    """API endpoint para obtener estadísticas"""
    reviews = load_reviews()
    stats = get_statistics(reviews)
    return jsonify(stats)

@app.route('/api/translate', methods=['POST'])
def api_translate():
    """API endpoint para traducir texto individual al alemán"""
    try:
        data = request.get_json()
        text = data.get('text', '')
        
        if not text:
            return jsonify({'error': 'No text provided'}), 400
        
        # Usar la función de traducción existente
        translated_text, was_translated, source_lang = translate_to_german(text)
        
        return jsonify({
            'original': text,
            'translated': translated_text,
            'was_translated': was_translated,
            'source_language': source_lang
        })
    
    except Exception as e:
        print(f"❌ Error en traducción API: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/analytics')
def analytics():
    """Página de análisis con gráficos temporales usando snapshots automáticos"""
    try:
        # Cargar metadata de snapshots automáticos
        metadata = load_snapshots_metadata()
        snapshot_list = metadata.get('snapshots', [])
        
        if not snapshot_list:
            return render_template('analytics.html', snapshots=[], has_data=False)
        
        # Procesar cada snapshot para generar estadísticas por apartamento
        processed_snapshots = []
        all_apartments = set()
        
        for snapshot_info in snapshot_list:
            snapshot_file = os.path.join(SNAPSHOTS_DIR, snapshot_info['filename'])
            
            if not os.path.exists(snapshot_file):
                continue
            
            # Leer archivo de snapshot
            with open(snapshot_file, 'r', encoding='utf-8') as f:
                reviews_data = json.load(f)
            
            # Calcular estadísticas por apartamento para este snapshot
            apartment_stats = {}
            
            for review in reviews_data:
                listing_url = review.get('listingUrl', '')
                rating = review.get('rating', 0)
                
                apartment_id = extract_apartment_id_from_url(listing_url)
                if not apartment_id or apartment_id in EXCLUDED_APARTMENT_IDS:
                    continue
                
                all_apartments.add(apartment_id)
                
                if apartment_id not in apartment_stats:
                    apartment_name = get_apartment_name_from_url(listing_url)
                    apartment_stats[apartment_id] = {
                        'apartment_id': apartment_id,
                        'apartment_name': apartment_name,
                        'total_reviews': 0,
                        'total_stars': 0,
                        'ratings': []
                    }
                
                if rating > 0:
                    apartment_stats[apartment_id]['total_reviews'] += 1
                    apartment_stats[apartment_id]['total_stars'] += rating
                    apartment_stats[apartment_id]['ratings'].append(rating)
            
            # Calcular promedios y porcentajes
            apartments_list = []
            for apt_id, stats in apartment_stats.items():
                if stats['total_reviews'] > 0:
                    avg_rating = stats['total_stars'] / stats['total_reviews']
                    five_stars = sum(1 for r in stats['ratings'] if r == 5)
                    five_star_pct = (five_stars / stats['total_reviews']) * 100 if stats['total_reviews'] > 0 else 0
                    
                    apartments_list.append({
                        'apartment_id': apt_id,
                        'apartment_name': stats['apartment_name'],
                        'total_reviews': stats['total_reviews'],
                        'average_rating': round(avg_rating, 2),
                        'five_star_percentage': round(five_star_pct, 1)
                    })
            
            # Agregar snapshot procesado
            processed_snapshots.append({
                'date': snapshot_info['created_at'][:10],  # Solo fecha (YYYY-MM-DD)
                'snapshot_id': snapshot_info['id'],
                'total_reviews': snapshot_info['total_reviews'],
                'apartments': apartments_list
            })
        
        # Preparar datos para cada apartamento (series temporales)
        apartments_data = {}
        for apt_id in all_apartments:
            apartments_data[apt_id] = {
                'name': '',
                'dates': [],
                'ratings': [],
                'reviews': [],
                'five_star_percentages': []
            }
        
        # Llenar datos por fecha
        for snapshot in processed_snapshots:
            date = snapshot['date']
            for apt in snapshot['apartments']:
                apt_id = apt['apartment_id']
                if apt_id in apartments_data:
                    apartments_data[apt_id]['name'] = apt['apartment_name']
                    apartments_data[apt_id]['dates'].append(date)
                    apartments_data[apt_id]['ratings'].append(apt['average_rating'])
                    apartments_data[apt_id]['reviews'].append(apt['total_reviews'])
                    apartments_data[apt_id]['five_star_percentages'].append(apt['five_star_percentage'])
        
        return render_template('analytics.html', 
                             snapshots=processed_snapshots,
                             apartments_data=apartments_data,
                             has_data=True)
    
    except Exception as e:
        print(f"❌ Error en analytics: {e}")
        import traceback
        traceback.print_exc()
        return render_template('analytics.html', snapshots=[], has_data=False)

@app.route('/rankings')
def rankings():
    """Página de rankings de apartamentos por estrellas (usa todas las reviews, sin filtro de fecha)"""
    try:
        # Leer todas las reviews sin filtro de fecha
        json_file_path = os.path.join(JSON_FOLDER_PATH, JSON_FILE_NAME)
        
        if not os.path.exists(json_file_path):
            return render_template('rankings.html', rankings=[], total_apartments=0, global_stats={})
        
        with open(json_file_path, 'r', encoding='utf-8') as file:
            all_reviews = json.load(file)
        
        # Calcular estadísticas por apartamento
        apartment_rankings = {}
        all_ratings = []  # Para estadísticas globales
        
        for review in all_reviews:
            listing_url = review.get('listingUrl', '')
            rating = review.get('rating', 0)
            
            # Extraer apartment_id
            apartment_id = extract_apartment_id_from_url(listing_url)
            if not apartment_id or apartment_id in EXCLUDED_APARTMENT_IDS:
                continue
            
            # Obtener nombre del apartamento
            apartment_name = get_apartment_name_from_url(listing_url)
            
            if apartment_id not in apartment_rankings:
                apartment_rankings[apartment_id] = {
                    'id': apartment_id,
                    'name': apartment_name,
                    'url': listing_url,
                    'total_reviews': 0,
                    'total_stars': 0,
                    'ratings': []
                }
            
            if rating > 0:
                apartment_rankings[apartment_id]['total_reviews'] += 1
                apartment_rankings[apartment_id]['total_stars'] += rating
                apartment_rankings[apartment_id]['ratings'].append(rating)
                all_ratings.append(rating)
        
        # Calcular promedio y distribución para cada apartamento
        rankings_list = []
        for apt_id, data in apartment_rankings.items():
            if data['total_reviews'] > 0:
                avg_rating = data['total_stars'] / data['total_reviews']
                
                # Distribución de estrellas
                rating_dist = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
                for r in data['ratings']:
                    if r in rating_dist:
                        rating_dist[r] += 1
                
                rankings_list.append({
                    'id': data['id'],
                    'name': data['name'],
                    'url': data['url'],
                    'average_rating': round(avg_rating, 2),
                    'total_reviews': data['total_reviews'],
                    'total_stars': data['total_stars'],
                    'rating_distribution': rating_dist,
                    'five_star_count': rating_dist[5],
                    'five_star_percentage': round((rating_dist[5] / data['total_reviews']) * 100) if data['total_reviews'] > 0 else 0
                })
        
        # Ordenar por promedio de estrellas (descendente)
        rankings_list.sort(key=lambda x: x['average_rating'], reverse=True)
        
        # Agrupar apartamentos por prefijo (completo)
        grouped_apartments = {}
        for apt in rankings_list:
            # Extraer prefijo completo (todas las letras antes del primer número)
            name = apt['name'] if apt['name'] else 'X'
            
            # Extraer prefijo: todas las letras consecutivas del inicio (sin números)
            prefix_match = re.match(r'^([A-Za-z]+)', name)
            if prefix_match:
                prefix = prefix_match.group(1).upper()  # Normalizar a mayúsculas
            else:
                prefix = 'OTHER'
            
            # Reglas especiales de agrupación
            if prefix == 'FAMZI':
                prefix = 'UO'  # FAMZI va con UO
            elif prefix == 'FO':
                prefix = 'HW'  # FO va con HW
            
            if prefix not in grouped_apartments:
                grouped_apartments[prefix] = {
                    'prefix': prefix,
                    'apartments': [],
                    'total_reviews': 0,
                    'total_stars': 0,
                    'count': 0
                }
            
            grouped_apartments[prefix]['apartments'].append(apt)
            grouped_apartments[prefix]['total_reviews'] += apt['total_reviews']
            grouped_apartments[prefix]['total_stars'] += apt['total_stars']
            grouped_apartments[prefix]['count'] += 1
        
        # Calcular promedio por grupo
        groups_list = []
        for prefix, group_data in grouped_apartments.items():
            avg_rating = group_data['total_stars'] / group_data['total_reviews'] if group_data['total_reviews'] > 0 else 0
            groups_list.append({
                'prefix': prefix,
                'apartments': group_data['apartments'],
                'average_rating': round(avg_rating, 2),
                'total_reviews': group_data['total_reviews'],
                'total_stars': group_data['total_stars'],
                'count': group_data['count']
            })
        
        # Ordenar grupos por promedio de estrellas
        groups_list.sort(key=lambda x: x['average_rating'], reverse=True)
        
        # Calcular estadísticas globales y destacados
        global_stats = {}
        if all_ratings and rankings_list:
            # Buscar apartamentos con más de cada tipo de estrella
            most_5_stars = max(rankings_list, key=lambda x: x['rating_distribution'][5])
            most_4_stars = max(rankings_list, key=lambda x: x['rating_distribution'][4])
            most_3_stars = max(rankings_list, key=lambda x: x['rating_distribution'][3])
            
            global_stats = {
                'total_reviews': len(all_ratings),
                'average_rating': round(sum(all_ratings) / len(all_ratings), 2),
                'five_star_total': sum(1 for r in all_ratings if r == 5),
                'four_star_total': sum(1 for r in all_ratings if r == 4),
                'three_star_total': sum(1 for r in all_ratings if r == 3),
                'rating_distribution': {
                    1: sum(1 for r in all_ratings if r == 1),
                    2: sum(1 for r in all_ratings if r == 2),
                    3: sum(1 for r in all_ratings if r == 3),
                    4: sum(1 for r in all_ratings if r == 4),
                    5: sum(1 for r in all_ratings if r == 5)
                },
                # Destacados (solo por tipo de estrellas)
                'most_5_stars': most_5_stars,
                'most_4_stars': most_4_stars,
                'most_3_stars': most_3_stars
            }
        
        return render_template('rankings.html',
                             rankings=rankings_list,
                             groups=groups_list,
                             total_apartments=len(rankings_list),
                             global_stats=global_stats)
    
    except Exception as e:
        print(f"❌ Error en rankings: {e}")
        import traceback
        traceback.print_exc()
        return render_template('rankings.html', rankings=[], total_apartments=0, global_stats={})

@app.route('/apartment-issues')
def apartment_issues():
    """Página de lista de problemas de apartamentos"""
    return render_template('apartment_issues.html')

@app.route('/api/apartment-issues')
def api_apartment_issues():
    """API endpoint para obtener problemas de apartamentos desde GeneralReviews.json"""
    try:
        # Buscar todos los archivos GeneralReviews con timestamp
        import glob
        general_files = glob.glob(os.path.join(JSON_FOLDER_PATH, "GeneralReviews_*.json"))
        
        # Si no hay archivos con timestamp, buscar el archivo legacy sin timestamp
        if not general_files:
            legacy_file = os.path.join(JSON_FOLDER_PATH, GENERAL_REVIEWS_FILE_NAME)
            if os.path.exists(legacy_file):
                general_files = [legacy_file]
        
        if not general_files:
            return jsonify({
                'error': 'GeneralReviews files not found',
                'wohnungen': []
            }), 404
        
        print(f"\n📁 Cargando GeneralReviews: {len(general_files)} archivo(s) encontrado(s)")
        
        # Ordenar archivos por nombre (más reciente primero)
        general_files.sort(reverse=True)

        # RECENTS: Solo el archivo más reciente
        most_recent_wohnungen = []
        if general_files:
            most_recent_file = general_files[0]
            try:
                with open(most_recent_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list) and len(data) > 0:
                        most_recent_wohnungen = data[0].get('message', {}).get('content', {}).get('wohnungen', [])
                        most_recent_wohnungen = normalize_probleme(most_recent_wohnungen)
                        print(f"   ✓ Más reciente: {os.path.basename(most_recent_file)}: {len(most_recent_wohnungen)} wohnungen")
            except Exception as e:
                print(f"   ⚠️  Error leyendo archivo más reciente: {e}")
        
        # Load history reviews from HRFinal.json
        hrfinal_path = os.path.join(os.path.dirname(__file__), 'HRFinal.json')
        history_wohnungen = []
        if os.path.exists(hrfinal_path):
            with open(hrfinal_path, 'r', encoding='utf-8') as f:
                hist_data = json.load(f)
            if isinstance(hist_data, list) and len(hist_data) > 0:
                history_wohnungen = hist_data[0].get('message', {}).get('content', {}).get('wohnungen', [])
                history_wohnungen = normalize_probleme(history_wohnungen)
        
        # Use only most recent file for recents
        recents_wohnungen = most_recent_wohnungen
        
        # Merge history and recents for general view
        merged = {}
        for apt in history_wohnungen:
            merged[apt.get('wohnung')] = {
                'wohnung': apt.get('wohnung'),
                'probleme': apt.get('probleme', []).copy()
            }
        for apt in recents_wohnungen:
            key = apt.get('wohnung')
            if key in merged:
                # Merge problems, avoid duplicates based on description
                existing = merged[key]['probleme']
                for prob in apt.get('probleme', []):
                    if not any(p.get('beschreibung') == prob.get('beschreibung') for p in existing):
                        existing.append(prob)
            else:
                merged[key] = {
                    'wohnung': apt.get('wohnung'),
                    'probleme': apt.get('probleme', []).copy()
                }
        general_wohnungen = list(merged.values())
        
        return jsonify({
            'history': history_wohnungen,
            'recents': recents_wohnungen,
            'general': general_wohnungen
        })
    except Exception as e:
        print(f"❌ Error al cargar problemas de apartamentos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': str(e),
            'wohnungen': []
        }), 500

@app.route('/api/apartment-reviews/<apartment_code>')
def api_apartment_reviews(apartment_code):
    """API endpoint para obtener reviews detalladas de un apartamento específico"""
    try:
        data_folder = os.path.join(os.path.dirname(__file__), 'data', 'DataProblemListing')
        
        # Map special apartment codes to their file names
        code_mapping = {
            'FZ': 'FAMZI',
            # Add more mappings here if needed
        }
        
        # Use mapped code if it exists, otherwise use original
        file_code = code_mapping.get(apartment_code, apartment_code)
        
        # Try both Airbnb and Booking files
        reviews = []
        
        # Load Airbnb reviews
        airbnb_file = os.path.join(data_folder, f'Airbnb{file_code}.json')
        if os.path.exists(airbnb_file):
            with open(airbnb_file, 'r', encoding='utf-8') as f:
                airbnb_data = json.load(f)
                for review in airbnb_data:
                    reviews.append({
                        'id': str(review.get('reviewId', '')),
                        'name': review.get('reviewerName', ''),
                        'date': review.get('reviewDate', ''),
                        'text': review.get('reviewText', ''),
                        'rating': review.get('rating', 0),
                        'source': 'Airbnb'
                    })
        
        # Load Booking reviews
        booking_file = os.path.join(data_folder, f'Booking{file_code}.json')
        if os.path.exists(booking_file):
            with open(booking_file, 'r', encoding='utf-8') as f:
                booking_data = json.load(f)
                for review in booking_data:
                    # Combine liked and disliked text
                    liked = review.get('likedText', '') or ''
                    disliked = review.get('dislikedText', '') or ''
                    full_text = ''
                    if liked:
                        full_text += f"👍 {liked}"
                    if disliked:
                        if full_text:
                            full_text += " | "
                        full_text += f"👎 {disliked}"
                    
                    reviews.append({
                        'id': '',  # Booking doesn't have review IDs
                        'name': review.get('userName', ''),
                        'date': review.get('reviewDate', ''),
                        'text': full_text or review.get('reviewTitle', ''),
                        'rating': review.get('rating', 0),
                        'source': 'Booking'
                    })
        
        return jsonify({
            'apartment': apartment_code,
            'reviews': reviews
        })
        
    except Exception as e:
        print(f"❌ Error al cargar reviews del apartamento {apartment_code}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': str(e),
            'reviews': []
        }), 500

@app.route('/api/current-reviews')
def api_current_reviews():
    """API endpoint para obtener reviews actuales desde DatasetScrAN.json y DatasetScrBookingAN.json"""
    try:
        import glob
        dataset_folder = r'C:\n8n_Docker\Files'

        reviews = []

        # Load Airbnb reviews from DatasetScrAN.json (main analyzed data source)
        airbnb_file = os.path.join(dataset_folder, 'DatasetScrAN.json')
        if os.path.exists(airbnb_file):
            try:
                with open(airbnb_file, 'r', encoding='utf-8') as f:
                    # Read as text first to preserve large numbers
                    import re
                    content = f.read()
                    # Convert large reviewId numbers to strings to prevent precision loss
                    # Match "reviewId": followed by a large number (15+ digits)
                    content = re.sub(r'"reviewId"\s*:\s*(\d{15,})', r'"reviewId": "\1"', content)
                    airbnb_data = json.loads(content)

                    if isinstance(airbnb_data, list):
                        for review in airbnb_data:
                            if 'reviewId' in review:
                                reviews.append({
                                    'id': str(review.get('reviewId', '')),
                                    'name': review.get('reviewerName', ''),
                                    'date': review.get('reviewDate', ''),
                                    'text': review.get('reviewText', ''),
                                    'rating': review.get('rating', 0),
                                    'source': 'Airbnb'
                                })
            except Exception as e:
                print(f"Warning: Error loading {airbnb_file}: {e}")
        
        # Load Booking reviews from DatasetScrBookingAN.json (updated data source)
        booking_file = os.path.join(dataset_folder, 'DatasetScrBookingAN.json')
        if os.path.exists(booking_file):
            with open(booking_file, 'r', encoding='utf-8') as f:
                booking_data = json.load(f)
                for review in booking_data:
                    # Combine liked and disliked text
                    liked = review.get('likedText', '') or ''
                    disliked = review.get('dislikedText', '') or ''
                    full_text = ''
                    if liked:
                        full_text += f"👍 {liked}"
                    if disliked:
                        if full_text:
                            full_text += " | "
                        full_text += f"👎 {disliked}"
                    
                    reviews.append({
                        'id': review.get('id', ''),
                        'name': review.get('userName', ''),
                        'date': review.get('reviewDate', ''),
                        'text': full_text or review.get('reviewTitle', ''),
                        'rating': review.get('rating', 0),
                        'source': 'Booking'
                    })
        
        return jsonify({
            'reviews': reviews,
            'total': len(reviews)
        })
        
    except Exception as e:
        print(f"❌ Error al cargar reviews actuales: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': str(e),
            'reviews': []
        }), 500

@app.route('/api/backup-dates')
def api_backup_dates():
    """API endpoint para listar meses de backup disponibles (archivos GeneralReviews con timestamp)"""
    try:
        import glob
        import re

        # Buscar archivos GeneralReviews con timestamp (formato: GeneralReviewsDD_MM_YY_HHMMSS.json)
        general_files = glob.glob(os.path.join(JSON_FOLDER_PATH, "GeneralReviews[0-9]*.json"))

        print(f"\n📂 Buscando backups de GeneralReviews: {len(general_files)} encontrados")

        # Nombres de meses en alemán
        month_names = {
            '01': 'Januar', '02': 'Februar', '03': 'März', '04': 'April',
            '05': 'Mai', '06': 'Juni', '07': 'Juli', '08': 'August',
            '09': 'September', '10': 'Oktober', '11': 'November', '12': 'Dezember'
        }

        # Agrupar por mes (MM_YY)
        months_found = {}
        for file_path in general_files:
            try:
                filename = os.path.basename(file_path)
                # Extraer fecha: GeneralReviewsDD_MM_YY_HHMMSS.json
                match = re.match(r'GeneralReviews(\d{2})_(\d{2})_(\d{2})_(\d+)\.json', filename)
                if match:
                    day, month, year, time_str = match.groups()
                    month_key = f"{month}_{year}"  # MM_YY

                    if month_key not in months_found:
                        full_year = f"20{year}"
                        month_name = month_names.get(month, month)
                        months_found[month_key] = {
                            'month_key': month_key,
                            'month_formatted': f"{month_name} {full_year}",
                            'month': month,
                            'year': year,
                            'files_count': 0
                        }
                    months_found[month_key]['files_count'] += 1
                    print(f"   ✓ {filename} -> {month_names.get(month, month)} 20{year}")
            except Exception as e:
                print(f"   ⚠️  Error parseando {filename}: {e}")
                continue

        # Ordenar por año y mes (más reciente primero)
        months_list = sorted(months_found.values(),
                           key=lambda x: (x['year'], x['month']),
                           reverse=True)

        return jsonify({'dates': months_list})

    except Exception as e:
        print(f"❌ Error al listar meses de backup: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'dates': []}), 500

@app.route('/api/backup-data/<month_key>')
def api_backup_data(month_key):
    """API endpoint para obtener datos de backup de un mes específico (MM_YY)"""
    try:
        import glob
        import re

        # Validar formato month_key: MM_YY
        parts = month_key.split('_')
        if len(parts) != 2 or len(parts[0]) != 2 or len(parts[1]) != 2:
            return jsonify({'error': 'Formato de month_key inválido (debe ser MM_YY)', 'wohnungen': []}), 400

        month = parts[0]
        year = parts[1]

        print(f"\n📂 Cargando backups del mes: {month}/20{year}")

        # Buscar todos los archivos GeneralReviews de ese mes
        # Formato: GeneralReviewsDD_MM_YY_HHMMSS.json
        pattern = os.path.join(JSON_FOLDER_PATH, f"GeneralReviews*_{month}_{year}_*.json")
        general_files = glob.glob(pattern)

        print(f"   📁 Archivos encontrados: {len(general_files)}")

        if not general_files:
            return jsonify({'error': 'Keine Backups für diesen Monat gefunden', 'wohnungen': []}), 404

        # Cargar y fusionar todos los archivos
        merged_wohnungen = {}  # Dict para fusionar por apartamento

        for file_path in general_files:
            try:
                filename = os.path.basename(file_path)
                print(f"   📄 Cargando: {filename}")

                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # Extract wohnungen from backup
                file_wohnungen = []
                if isinstance(data, list) and len(data) > 0:
                    file_wohnungen = data[0].get('message', {}).get('content', {}).get('wohnungen', [])
                    file_wohnungen = normalize_probleme(file_wohnungen)

                # Fusionar con datos existentes
                for wohnung in file_wohnungen:
                    apt_name = wohnung.get('wohnung', 'Unknown')
                    if apt_name not in merged_wohnungen:
                        merged_wohnungen[apt_name] = {
                            'wohnung': apt_name,
                            'probleme': {}
                        }

                    # Fusionar problemas
                    for problem in wohnung.get('probleme', []):
                        desc = problem.get('beschreibung', '')
                        if not desc:
                            continue

                        if desc not in merged_wohnungen[apt_name]['probleme']:
                            merged_wohnungen[apt_name]['probleme'][desc] = {
                                'beschreibung': desc,
                                'erwähnungen': 0,
                                'kategorie': problem.get('kategorie', 'Sonstiges'),
                                'ids': [],
                                'names': []
                            }

                        # Sumar erwähnungen
                        erw = problem.get('erwähnungen') or problem.get('erw??hnungen') or 0
                        merged_wohnungen[apt_name]['probleme'][desc]['erwähnungen'] += erw

                        # Fusionar ids
                        prob_ids = problem.get('ids', []) or problem.get('id', [])
                        if isinstance(prob_ids, str):
                            prob_ids = [prob_ids]
                        elif not isinstance(prob_ids, list):
                            prob_ids = [str(prob_ids)] if prob_ids else []
                        for pid in prob_ids:
                            if pid and str(pid) not in merged_wohnungen[apt_name]['probleme'][desc]['ids']:
                                merged_wohnungen[apt_name]['probleme'][desc]['ids'].append(str(pid))

                        # Fusionar names
                        prob_names = problem.get('names', [])
                        if isinstance(prob_names, str):
                            prob_names = [prob_names]
                        for name in prob_names:
                            if name and name not in merged_wohnungen[apt_name]['probleme'][desc]['names']:
                                merged_wohnungen[apt_name]['probleme'][desc]['names'].append(name)

                print(f"      ✓ {len(file_wohnungen)} wohnungen procesadas")

            except Exception as e:
                print(f"   ⚠️  Error cargando {filename}: {e}")
                continue

        # Convertir dict a lista
        result_wohnungen = []
        for apt_name, apt_data in merged_wohnungen.items():
            result_wohnungen.append({
                'wohnung': apt_name,
                'probleme': list(apt_data['probleme'].values())
            })

        # Ordenar por nombre de apartamento
        result_wohnungen.sort(key=lambda x: x['wohnung'])

        print(f"   ✅ Total: {len(result_wohnungen)} wohnungen con {sum(len(w['probleme']) for w in result_wohnungen)} problemas")

        return jsonify({
            'month_key': month_key,
            'wohnungen': result_wohnungen,
            'files_loaded': len(general_files)
        })

    except Exception as e:
        print(f"❌ Error al cargar backup {month_key}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': str(e),
            'wohnungen': []
        }), 500

@app.route('/api/backup-reviews/<month_key>')
def api_backup_reviews(month_key):
    """API endpoint para obtener reviews de backup de un mes específico (MM_YY)"""
    try:
        import glob

        # Validar formato month_key: MM_YY
        parts = month_key.split('_')
        if len(parts) != 2 or len(parts[0]) != 2 or len(parts[1]) != 2:
            return jsonify({'error': 'Formato de month_key inválido (debe ser MM_YY)', 'reviews': []}), 400

        month = parts[0]
        year = parts[1]

        print(f"\n📂 Cargando reviews del mes: {month}/20{year}")

        reviews = []
        seen_ids = set()  # Para evitar duplicados

        # Buscar archivos Airbnb de ese mes
        # Formato: DatasetScr_DD_MM_YY_HHMMSS.json
        airbnb_pattern = os.path.join(JSON_FOLDER_PATH, f"DatasetScr_*_{month}_{year}_*.json")
        airbnb_files = glob.glob(airbnb_pattern)

        print(f"   📁 Archivos Airbnb encontrados: {len(airbnb_files)}")

        for file_path in airbnb_files:
            try:
                filename = os.path.basename(file_path)
                with open(file_path, 'r', encoding='utf-8') as f:
                    airbnb_data = json.load(f)
                    for review in airbnb_data:
                        review_id = str(review.get('reviewId', ''))
                        unique_key = f"airbnb_{review_id}"
                        if unique_key not in seen_ids:
                            seen_ids.add(unique_key)
                            reviews.append({
                                'id': review_id,
                                'name': review.get('reviewerName', ''),
                                'date': review.get('reviewDate', ''),
                                'text': review.get('reviewText', ''),
                                'rating': review.get('rating', 0),
                                'source': 'Airbnb'
                            })
                print(f"      ✓ {filename}: {len(airbnb_data)} reviews")
            except Exception as e:
                print(f"      ⚠️  Error cargando {filename}: {e}")

        # Buscar archivos Booking de ese mes
        # Formato: DatasetScrBooking_DD_MM_YY_HHMMSS.json
        booking_pattern = os.path.join(JSON_FOLDER_PATH, f"DatasetScrBooking_*_{month}_{year}_*.json")
        booking_files = glob.glob(booking_pattern)

        print(f"   📁 Archivos Booking encontrados: {len(booking_files)}")

        for file_path in booking_files:
            try:
                filename = os.path.basename(file_path)
                with open(file_path, 'r', encoding='utf-8') as f:
                    booking_data = json.load(f)
                    for review in booking_data:
                        reviewer_name = review.get('userName', '')
                        review_date = review.get('reviewDate', '')
                        unique_key = f"booking_{reviewer_name}_{review_date}"
                        if unique_key not in seen_ids:
                            seen_ids.add(unique_key)
                            # Combine liked and disliked text
                            liked = review.get('likedText', '') or ''
                            disliked = review.get('dislikedText', '') or ''
                            full_text = ''
                            if liked:
                                full_text += f"👍 {liked}"
                            if disliked:
                                if full_text:
                                    full_text += " | "
                                full_text += f"👎 {disliked}"

                            reviews.append({
                                'id': review.get('id', ''),
                                'name': reviewer_name,
                                'date': review_date,
                                'text': full_text or review.get('reviewTitle', ''),
                                'rating': review.get('rating', 0),
                                'source': 'Booking'
                            })
                print(f"      ✓ {filename}: {len(booking_data)} reviews")
            except Exception as e:
                print(f"      ⚠️  Error cargando {filename}: {e}")

        print(f"   ✅ Total: {len(reviews)} reviews únicas")

        return jsonify({
            'month_key': month_key,
            'reviews': reviews,
            'total': len(reviews)
        })

    except Exception as e:
        print(f"❌ Error al cargar reviews de backup {month_key}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': str(e),
            'reviews': []
        }), 500

@app.route('/api/snapshots')
def api_list_snapshots():
    """API endpoint para listar todos los snapshots disponibles"""
    try:
        metadata = load_snapshots_metadata()
        return jsonify(metadata)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/create-snapshot', methods=['POST'])
def api_create_snapshot():
    """API endpoint para crear un snapshot manual"""
    try:
        json_file_path = os.path.join(JSON_FOLDER_PATH, JSON_FILE_NAME)
        
        if not os.path.exists(json_file_path):
            return jsonify({'error': 'Dataset no encontrado'}), 404
        
        snapshot_info = create_snapshot(json_file_path)
        
        if snapshot_info:
            # Actualizar hash
            current_hash = calculate_file_hash(json_file_path)
            save_hash(current_hash)
            
            return jsonify({
                'success': True,
                'snapshot': snapshot_info,
                'message': 'Snapshot creado exitosamente'
            })
        else:
            return jsonify({'error': 'No se pudo crear el snapshot'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics-weekly')
def api_analytics_weekly():
    """API endpoint para analytics semanal usando BackupIssues"""
    try:
        backup_folder = os.path.join(os.path.dirname(__file__), 'data', 'BackupIssues')
        
        if not os.path.exists(backup_folder):
            return jsonify({'error': 'BackupIssues folder not found', 'backups': []}), 404
        
        # Buscar todas las fechas de backup (desde DatasetScr)
        files = os.listdir(backup_folder)
        backup_dates = set()
        
        for filename in files:
            if filename.startswith('DatasetScr') and filename.endswith('.json') and not filename.startswith('DatasetScrBooking'):
                # Extraer fecha: DatasetScr021225.json -> 021225
                date_part = filename.replace('DatasetScr', '').replace('.json', '')
                if len(date_part) == 6 and date_part.isdigit():
                    backup_dates.add(date_part)
        
        if not backup_dates:
            return jsonify({'error': 'No backups available', 'backups': []}), 404
        
        # Ordenar fechas (más antiguo primero para timeline)
        sorted_dates = sorted(list(backup_dates))
        
        # Procesar cada backup
        processed_backups = []
        all_apartments = {}
        
        for date_key in sorted_dates:
            # Convertir fecha DDMMYY a formato legible
            try:
                from datetime import datetime
                day = date_key[0:2]
                month = date_key[2:4]
                year = '20' + date_key[4:6]
                date_obj = datetime.strptime(f"{day}{month}{year}", "%d%m%Y")
                date_formatted = date_obj.strftime('%Y-%m-%d')
            except:
                continue
            
            # Leer archivos de este backup
            airbnb_file = os.path.join(backup_folder, f'DatasetScr{date_key}.json')
            booking_file = os.path.join(backup_folder, f'DatasetScrBooking{date_key}.json')
            
            apartment_stats = {}
            
            # Procesar Airbnb
            if os.path.exists(airbnb_file):
                with open(airbnb_file, 'r', encoding='utf-8') as f:
                    airbnb_reviews = json.load(f)
                
                for review in airbnb_reviews:
                    listing_url = review.get('listingUrl', '')
                    apartment_id = extract_apartment_id_from_url(listing_url)
                    
                    if not apartment_id or apartment_id in EXCLUDED_APARTMENT_IDS:
                        continue
                    
                    apt_code = get_apartment_name_from_url(listing_url)
                    rating = review.get('rating', 0)
                    
                    if apt_code not in apartment_stats:
                        apartment_stats[apt_code] = {'airbnb_ratings': [], 'booking_ratings': []}
                    
                    if rating > 0:
                        apartment_stats[apt_code]['airbnb_ratings'].append(rating)
            
            # Procesar Booking
            if os.path.exists(booking_file):
                with open(booking_file, 'r', encoding='utf-8') as f:
                    booking_reviews = json.load(f)
                
                for review in booking_reviews:
                    booking_url = review.get('startUrl', '')
                    apt_code = get_apartment_code_from_booking_url(booking_url)
                    
                    if not apt_code:
                        continue
                    
                    rating = review.get('rating', 0)
                    
                    if apt_code not in apartment_stats:
                        apartment_stats[apt_code] = {'airbnb_ratings': [], 'booking_ratings': []}
                    
                    if rating > 0:
                        apartment_stats[apt_code]['booking_ratings'].append(rating)
            
            # Calcular promedios combinados (normalizar Booking a escala 5)
            apartments_list = []
            for apt_code, stats in apartment_stats.items():
                airbnb_ratings = stats['airbnb_ratings']
                booking_ratings = stats['booking_ratings']
                
                # Normalizar Booking de 10 a 5
                all_ratings_normalized = airbnb_ratings + [r / 2 for r in booking_ratings]
                
                if all_ratings_normalized:
                    avg_rating = sum(all_ratings_normalized) / len(all_ratings_normalized)
                    apartments_list.append({
                        'code': apt_code,
                        'average_rating': round(avg_rating, 2),
                        'total_reviews': len(all_ratings_normalized)
                    })
                    
                    # Registrar apartamento
                    if apt_code not in all_apartments:
                        all_apartments[apt_code] = {'dates': [], 'ratings': []}
                    
                    all_apartments[apt_code]['dates'].append(date_formatted)
                    all_apartments[apt_code]['ratings'].append(round(avg_rating, 2))
            
            processed_backups.append({
                'date': date_formatted,
                'date_key': date_key,
                'apartments': apartments_list
            })
        
        # Preparar timeline
        timeline = []
        for apt_code, data in all_apartments.items():
            timeline.append({
                'code': apt_code,
                'dates': data['dates'],
                'ratings': data['ratings']
            })
        
        # Calcular tendencias
        trends = []
        for apt_code, data in all_apartments.items():
            if len(data['ratings']) >= 2:
                first_rating = data['ratings'][0]
                last_rating = data['ratings'][-1]
                change = last_rating - first_rating
                change_pct = (change / first_rating) * 100 if first_rating > 0 else 0
                
                trends.append({
                    'code': apt_code,
                    'first_rating': round(first_rating, 2),
                    'last_rating': round(last_rating, 2),
                    'change': round(change, 2),
                    'change_percentage': round(change_pct, 1),
                    'trend': 'up' if change > 0 else ('down' if change < 0 else 'stable')
                })
        
        # Ordenar trends por cambio (descendente)
        trends.sort(key=lambda x: x['change'], reverse=True)
        
        return jsonify({
            'timeline': timeline,
            'trends': trends,
            'backup_count': len(processed_backups),
            'date_range': {
                'start': processed_backups[0]['date'] if processed_backups else None,
                'end': processed_backups[-1]['date'] if processed_backups else None
            }
        })
        
    except Exception as e:
        print(f"❌ Error en api_analytics_weekly: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics-general')
def api_analytics_general():
    """API endpoint para analytics usando datos históricos de DataProblemListing"""
    try:
        data_folder = os.path.join(os.path.dirname(__file__), 'data', 'DataProblemListing')
        
        if not os.path.exists(data_folder):
            return jsonify({'error': 'DataProblemListing folder not found'}), 404
        
        # Diccionario para almacenar stats por apartamento
        apartments = {}
        
        # Leer todos los archivos JSON
        for filename in os.listdir(data_folder):
            if not filename.endswith('.json'):
                continue
            
            # Determinar fuente (Airbnb o Booking) y código de apartamento
            if filename.startswith('Airbnb'):
                source = 'Airbnb'
                apt_code = filename.replace('Airbnb', '').replace('.json', '')
            elif filename.startswith('Booking'):
                source = 'Booking'
                apt_code = filename.replace('Booking', '').replace('.json', '')
            else:
                continue
            
            # Leer archivo
            file_path = os.path.join(data_folder, filename)
            with open(file_path, 'r', encoding='utf-8') as f:
                reviews = json.load(f)
            
            # Inicializar apartamento si no existe
            if apt_code not in apartments:
                apartments[apt_code] = {
                    'code': apt_code,
                    'airbnb': {'reviews': [], 'ratings': [], 'dates': []},
                    'booking': {'reviews': [], 'ratings': [], 'dates': []}
                }
            
            # Procesar reviews
            for review in reviews:
                if source == 'Airbnb':
                    rating = review.get('rating', 0)
                    date = review.get('reviewDate', '')
                    if rating > 0:
                        apartments[apt_code]['airbnb']['reviews'].append(review)
                        apartments[apt_code]['airbnb']['ratings'].append(rating)
                        apartments[apt_code]['airbnb']['dates'].append(date)
                else:  # Booking
                    rating = review.get('rating', 0)
                    date = review.get('reviewDate', '')
                    if rating > 0:
                        apartments[apt_code]['booking']['reviews'].append(review)
                        apartments[apt_code]['booking']['ratings'].append(rating)
                        apartments[apt_code]['booking']['dates'].append(date)
        
        # Calcular estadísticas por apartamento
        result = []
        for apt_code, data in apartments.items():
            airbnb_ratings = data['airbnb']['ratings']
            booking_ratings = data['booking']['ratings']
            
            apt_stats = {
                'code': apt_code,
                'airbnb': {
                    'total_reviews': len(airbnb_ratings),
                    'average_rating': round(sum(airbnb_ratings) / len(airbnb_ratings), 2) if airbnb_ratings else 0,
                    'five_star_count': sum(1 for r in airbnb_ratings if r == 5),
                    'five_star_percentage': round((sum(1 for r in airbnb_ratings if r == 5) / len(airbnb_ratings)) * 100, 1) if airbnb_ratings else 0
                },
                'booking': {
                    'total_reviews': len(booking_ratings),
                    'average_rating': round(sum(booking_ratings) / len(booking_ratings), 2) if booking_ratings else 0,
                    'ten_rating_count': sum(1 for r in booking_ratings if r == 10),
                    'ten_rating_percentage': round((sum(1 for r in booking_ratings if r == 10) / len(booking_ratings)) * 100, 1) if booking_ratings else 0
                },
                'combined': {
                    'total_reviews': len(airbnb_ratings) + len(booking_ratings)
                }
            }
            
            # Calcular rating combinado (normalizar Booking de 10 a 5)
            all_ratings_normalized = airbnb_ratings + [r / 2 for r in booking_ratings]
            if all_ratings_normalized:
                apt_stats['combined']['average_rating'] = round(sum(all_ratings_normalized) / len(all_ratings_normalized), 2)
            else:
                apt_stats['combined']['average_rating'] = 0
            
            result.append(apt_stats)
        
        # Ordenar por código de apartamento
        result.sort(key=lambda x: x['code'])
        
        return jsonify({
            'apartments': result,
            'total_apartments': len(result)
        })
        
    except Exception as e:
        print(f"❌ Error en api_analytics_general: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def preload_cache():
    """Pre-cargar el caché de reviews al iniciar el servidor"""
    print("\n🚀 Pre-cargando caché de reviews...")
    try:
        load_reviews()
        print("✅ Caché de reviews nuevas cargado")
        load_reviews_for_slideshow()
        print("✅ Caché de slideshow (con traducciones) cargado")
    except Exception as e:
        print(f"⚠️  Error pre-cargando caché: {e}")

if __name__ == '__main__':
    # Log crashes to file so we can debug when the console window closes
    import logging
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler('server.log', encoding='utf-8')
        ]
    )

    # Pre-cargar caché antes de iniciar el servidor
    preload_cache()

    # Production WSGI server (handles concurrent requests properly)
    from waitress import serve
    print("Starting Waitress server on http://0.0.0.0:80 (16 threads)...")
    serve(app, host='0.0.0.0', port=80, threads=16)
