import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_excel_professional():
    print("=" * 80)
    print("FORMATEANDO EXCEL CON ESTILO PROFESIONAL")
    print("=" * 80)
    print()
    
    # Leer el Excel enriquecido
    input_file = Path(r'C:\Users\admin\Server\FlaskApp\positive_reviews_enriched.xlsx')
    output_file = Path(r'C:\Users\admin\Server\FlaskApp\positive_reviews_FINAL.xlsx')
    
    print("📖 Leyendo archivo...")
    df = pd.read_excel(input_file)
    print(f"   ✓ {len(df)} registros cargados")
    print()
    
    # Reordenar columnas para mejor lectura
    print("🔄 Reordenando columnas...")
    column_order = [
        'Plattform',
        'Wohnung', 
        'Name',
        'Nombre_Completo',
        'Email',
        'Telefon',
        'Direccion',
        'Datum',
        'Originalsprache',
        'Gefundenes Muster',
        'Kommentar'
    ]
    df = df[column_order]
    print("   ✓ Columnas reordenadas")
    print()
    
    # Guardar temporalmente
    print("💾 Guardando archivo...")
    df.to_excel(output_file, index=False, engine='openpyxl')
    print("   ✓ Archivo guardado")
    print()
    
    # Cargar workbook para aplicar estilos
    print("🎨 Aplicando estilos profesionales...")
    wb = load_workbook(output_file)
    ws = wb.active
    ws.title = "Reviews Positivas"
    
    # Definir estilos
    # Header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
    
    # Colores alternados por plataforma
    airbnb_fill = PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid")  # Rosa claro
    booking_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")  # Azul claro
    
    # Fill para contacto disponible
    contact_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Aplicar estilo a encabezados
    print("   → Formateando encabezados...")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Ajustar anchos de columna
    print("   → Ajustando anchos de columnas...")
    column_widths = {
        'A': 12,  # Plattform
        'B': 10,  # Wohnung
        'C': 18,  # Name
        'D': 25,  # Nombre_Completo
        'E': 28,  # Email
        'F': 18,  # Telefon
        'G': 35,  # Direccion
        'H': 12,  # Datum
        'I': 12,  # Originalsprache
        'J': 25,  # Gefundenes Muster
        'K': 55   # Kommentar
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Aplicar estilos a las filas de datos
    print("   → Aplicando colores por plataforma y contacto...")
    for row_idx in range(2, ws.max_row + 1):
        platform = ws[f'A{row_idx}'].value
        nombre_completo = ws[f'D{row_idx}'].value
        email = ws[f'E{row_idx}'].value
        telefon = ws[f'F{row_idx}'].value
        direccion = ws[f'G{row_idx}'].value
        
        # Determinar color de fondo según plataforma
        if platform == 'Airbnb':
            row_fill = airbnb_fill
        else:
            row_fill = booking_fill
        
        # Aplicar formato a cada celda de la fila
        for col_idx in range(1, 12):  # 11 columnas
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx == 11))  # Wrap en Kommentar
            
            # Color de fondo
            cell.fill = row_fill
            
            # Destacar celdas con contacto/info adicional
            if col_idx == 4 and nombre_completo:  # Nombre_Completo
                cell.fill = contact_fill
                cell.font = Font(bold=True, color="006100")
            elif col_idx == 5 and email:  # Email
                cell.fill = contact_fill
                cell.font = Font(bold=True, color="006100")
            elif col_idx == 6 and telefon:  # Telefon
                cell.fill = contact_fill
                cell.font = Font(bold=True, color="006100")
            elif col_idx == 7 and direccion:  # Direccion
                cell.fill = contact_fill
                cell.font = Font(bold=True, color="006100")
    
    # Ajustar altura de filas
    print("   → Ajustando altura de filas...")
    ws.row_dimensions[1].height = 30  # Header más alto
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 40  # Más espacio para comentarios
    
    # Congelar primera fila y primeras 3 columnas
    print("   → Congelando paneles...")
    ws.freeze_panes = 'D2'  # Congela hasta la columna C (Plattform, Wohnung, Name) y fila 1
    
    # Agregar autofiltro
    print("   → Agregando filtros...")
    ws.auto_filter.ref = ws.dimensions
    
    # Crear hoja de resumen
    print("   → Creando hoja de resumen...")
    ws_summary = wb.create_sheet("Resumen", 0)
    
    # Título de resumen
    ws_summary['A1'] = 'RESUMEN DE REVIEWS POSITIVAS'
    ws_summary['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws_summary['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Estadísticas
    total_reviews = len(df)
    reviews_with_full_name = df['Nombre_Completo'].notna().sum()
    reviews_with_email = df['Email'].notna().sum()
    reviews_with_phone = df['Telefon'].notna().sum()
    reviews_with_address = df['Direccion'].notna().sum()
    reviews_with_contact = df[(df['Email'].notna()) | (df['Telefon'].notna())].shape[0]
    
    stats = [
        ['', ''],
        ['Métrica', 'Valor'],
        ['Total de Reviews', total_reviews],
        ['Reviews con Nombre Completo', f"{reviews_with_full_name} ({reviews_with_full_name/total_reviews*100:.1f}%)"],
        ['Reviews con Email', f"{reviews_with_email} ({reviews_with_email/total_reviews*100:.1f}%)"],
        ['Reviews con Teléfono', f"{reviews_with_phone} ({reviews_with_phone/total_reviews*100:.1f}%)"],
        ['Reviews con Dirección', f"{reviews_with_address} ({reviews_with_address/total_reviews*100:.1f}%)"],
        ['Reviews con Algún Contacto', f"{reviews_with_contact} ({reviews_with_contact/total_reviews*100:.1f}%)"],
        ['', ''],
        ['Por Plataforma', ''],
        ['Airbnb', len(df[df['Plattform'] == 'Airbnb'])],
        ['Booking.com', len(df[df['Plattform'] == 'Booking.com'])],
        ['', ''],
        ['Top 5 Apartamentos', 'Reviews'],
    ]
    
    top_apartments = df['Wohnung'].value_counts().head(5)
    for apt, count in top_apartments.items():
        stats.append([apt, count])
    
    for row_idx, (label, value) in enumerate(stats, start=1):
        ws_summary[f'A{row_idx}'] = label
        ws_summary[f'B{row_idx}'] = value
        
        if row_idx == 2 or row_idx == 8 or row_idx == 12:  # Headers
            ws_summary[f'A{row_idx}'].font = Font(bold=True, color="FFFFFF")
            ws_summary[f'B{row_idx}'].font = Font(bold=True, color="FFFFFF")
            ws_summary[f'A{row_idx}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws_summary[f'B{row_idx}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # Guardar archivo final
    print("   ✓ Estilos aplicados")
    print()
    
    print("💾 Guardando archivo final...")
    wb.save(output_file)
    print(f"   ✓ Archivo guardado: {output_file.name}")
    print()
    
    # Resumen final
    print("=" * 80)
    print("✅ EXCEL PROFESIONAL GENERADO")
    print("=" * 80)
    print()
    print("📁 Archivo: positive_reviews_FINAL.xlsx")
    print()
    print("🎨 Características:")
    print("   ✓ Hoja de Resumen con estadísticas")
    print("   ✓ Encabezados azules con texto blanco")
    print("   ✓ Colores alternados por plataforma:")
    print("     • Airbnb: Rosa claro")
    print("     • Booking: Azul claro")
    print("   ✓ Emails y teléfonos destacados en verde")
    print("   ✓ Columnas organizadas lógicamente")
    print("   ✓ Filtros automáticos")
    print("   ✓ Primeras columnas congeladas")
    print("   ✓ Altura de filas ajustada")
    print("   ✓ Texto envuelto en comentarios")
    print()
    print("=" * 80)

if __name__ == '__main__':
    format_excel_professional()
