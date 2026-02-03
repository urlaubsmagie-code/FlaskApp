import json
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_to_excel():
    # Leer el archivo JSON generado
    json_file = Path(r'C:\Users\admin\Server\FlaskApp\positive_reviews_report.json')
    
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reviews Positivas"
    
    # Estilos
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    airbnb_fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
    booking_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados en alemán
    headers = ['Plattform', 'Wohnung', 'Name', 'Datum', 'Kommentar', 'Gefundenes Muster', 'Originalsprache']
    ws.append(headers)
    
    # Estilo de encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Agregar reviews de Airbnb
    for review in data['reviews']['airbnb']:
        # Formatear fecha
        fecha = review['reviewDate'][:10] if review['reviewDate'] else ''
        
        # Usar texto traducido si existe, sino el original
        comentario = review.get('translated_text', review['full_text']).replace('<br/>', ' ').replace('<br>', ' ').strip()
        if len(comentario) > 500:  # Limitar longitud
            comentario = comentario[:500] + '...'
        
        # Obtener idioma original
        idioma_original = review.get('original_language', review.get('language', 'unknown'))
        
        # Patrones encontrados
        patrones = ', '.join([p.replace('\\s+', ' ').replace('?', '') for p in review['matched_patterns']])
        
        row = [
            'Airbnb',
            review.get('apartmentCode', 'UNKNOWN'),
            review['reviewerName'],
            fecha,
            comentario,
            patrones,
            idioma_original
        ]
        ws.append(row)
        
        # Aplicar estilo de fila
        current_row = ws.max_row
        for cell in ws[current_row]:
            cell.fill = airbnb_fill
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Agregar reviews de Booking
    for review in data['reviews']['booking']:
        # Formatear fecha
        fecha = review['reviewDate'][:10] if review['reviewDate'] else ''
        
        # Usar texto traducido si existe
        if review.get('translated_text'):
            comentario = review['translated_text']
        else:
            # Combinar liked y disliked
            comentario = ''
            if review['likedText']:
                comentario += f"👍 {review['likedText']} "
            if review['dislikedText']:
                comentario += f"👎 {review['dislikedText']}"
            comentario = comentario.strip()
        
        if len(comentario) > 500:  # Limitar longitud
            comentario = comentario[:500] + '...'
        
        # Obtener idioma original
        idioma_original = review.get('original_language', review.get('reviewLanguage', 'unknown'))
        
        # Patrones encontrados
        patrones = ', '.join([p.replace('\\s+', ' ').replace('?', '') for p in review['matched_patterns']])
        
        row = [
            'Booking.com',
            review.get('apartmentCode', 'UNKNOWN'),
            review['userName'],
            fecha,
            comentario,
            patrones,
            idioma_original
        ]
        ws.append(row)
        
        # Aplicar estilo de fila
        current_row = ws.max_row
        for cell in ws[current_row]:
            cell.fill = booking_fill
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15  # Plataforma
    ws.column_dimensions['B'].width = 12  # Wohnung
    ws.column_dimensions['C'].width = 25  # Nombre
    ws.column_dimensions['D'].width = 12  # Fecha
    ws.column_dimensions['E'].width = 80  # Comentario
    ws.column_dimensions['F'].width = 25  # Patrón
    ws.column_dimensions['G'].width = 12  # Originalsprache
    
    # Ajustar altura de filas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 50
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Guardar archivo
    excel_file = Path(r'C:\Users\admin\Server\FlaskApp\positive_reviews_simple.xlsx')
    wb.save(excel_file)
    
    print(f"✓ Archivo Excel generado exitosamente: {excel_file}")
    print(f"\nTotal de reviews exportadas: {data['total_positive_reviews']}")
    print(f"  - Airbnb: {data['airbnb_count']}")
    print(f"  - Booking: {data['booking_count']}")
    print(f"\n💡 Puedes abrir este archivo en Excel, Google Sheets, LibreOffice, etc.")

if __name__ == '__main__':
    export_to_excel()
